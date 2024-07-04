import json
import logging
from datetime import datetime, time
from io import BytesIO

import pandas as pd
from django.core.cache import cache
from django.db import models, transaction
from django.db.models import CharField, Q, Value
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from ledger_api_client.ledger_models import EmailUserRO as EmailUser
from multiselectfield import MultiSelectField
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import mixins, serializers, status, views, viewsets
from rest_framework.decorators import action as detail_route
from rest_framework.decorators import action as list_route
from rest_framework.decorators import renderer_classes
from rest_framework.exceptions import PermissionDenied
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework_datatables.filters import DatatablesFilterBackend
from rest_framework_datatables.pagination import DatatablesPageNumberPagination
from django.contrib.gis.geos import GEOSGeometry

from boranga import settings
from boranga.components.conservation_status.serializers import SendReferralSerializer
from boranga.components.main.api import search_datums
from boranga.components.main.related_item import RelatedItemsSerializer
from boranga.components.main.utils import validate_threat_request
from boranga.components.occurrence.email import send_external_referee_invite_email
from boranga.components.occurrence.filters import OccurrenceReportReferralFilterBackend
from boranga.components.occurrence.mixins import DatumSearchMixin
from boranga.components.occurrence.models import (
    AnimalHealth,
    CoordinateSource,
    CountedSubject,
    Datum,
    DeathReason,
    Drainage,
    IdentificationCertainty,
    Intensity,
    LandForm,
    LocationAccuracy,
    ObservationMethod,
    OCCAnimalObservation,
    OCCAssociatedSpecies,
    OCCConservationThreat,
    OCCContactDetail,
    OCCFireHistory,
    OCCHabitatComposition,
    OCCHabitatCondition,
    OCCIdentification,
    OCCLocation,
    OCCObservationDetail,
    OCCPlantCount,
    Occurrence,
    OccurrenceDocument,
    OccurrenceGeometry,
    OccurrenceReport,
    OccurrenceReportAmendmentRequest,
    OccurrenceReportAmendmentRequestDocument,
    OccurrenceReportDocument,
    OccurrenceReportGeometry,
    OccurrenceReportReferral,
    OccurrenceReportUserAction,
    OccurrenceSite,
    OccurrenceTenure,
    OccurrenceUserAction,
    OCCVegetationStructure,
    OCRAnimalObservation,
    OCRAssociatedSpecies,
    OCRConservationThreat,
    OCRExternalRefereeInvite,
    OCRFireHistory,
    OCRHabitatComposition,
    OCRHabitatCondition,
    OCRIdentification,
    OCRLocation,
    OCRObservationDetail,
    OCRObserverDetail,
    OCRPlantCount,
    OCRVegetationStructure,
    PermitType,
    PlantCondition,
    PlantCountAccuracy,
    PlantCountMethod,
    PrimaryDetectionMethod,
    ReproductiveState,
    RockType,
    SampleDestination,
    SampleType,
    SecondarySign,
    SiteType,
    SoilColour,
    SoilCondition,
    SoilType,
    WildStatus,
)
from boranga.components.occurrence.serializers import (
    BackToAssessorSerializer,
    CreateOccurrenceReportSerializer,
    CreateOccurrenceSerializer,
    DTOccurrenceReportReferralSerializer,
    InternalOccurrenceReportSerializer,
    InternalSaveOccurrenceReportDocumentSerializer,
    ListInternalOccurrenceReportSerializer,
    ListOCCMinimalSerializer,
    ListOccurrenceReportSerializer,
    ListOccurrenceSerializer,
    ListOccurrenceTenureSerializer,
    ListOCRReportMinimalSerializer,
    OCCConservationThreatSerializer,
    OCCContactDetailSerializer,
    OccurrenceDocumentSerializer,
    OccurrenceLogEntrySerializer,
    OccurrenceReportAmendmentRequestSerializer,
    OccurrenceReportDocumentSerializer,
    OccurrenceReportLogEntrySerializer,
    OccurrenceReportProposalReferralSerializer,
    OccurrenceReportReferralSerializer,
    OccurrenceReportSerializer,
    OccurrenceReportUserActionSerializer,
    OccurrenceSerializer,
    OccurrenceSiteSerializer,
    OccurrenceTenureSerializer,
    OccurrenceUserActionSerializer,
    OCRConservationThreatSerializer,
    OCRExternalRefereeInviteSerializer,
    OCRObserverDetailLimitedSerializer,
    OCRObserverDetailSerializer,
    ProposeApproveSerializer,
    ProposeDeclineSerializer,
    SaveOCCAnimalObservationSerializer,
    SaveOCCAssociatedSpeciesSerializer,
    SaveOCCConservationThreatSerializer,
    SaveOCCFireHistorySerializer,
    SaveOCCHabitatCompositionSerializer,
    SaveOCCHabitatConditionSerializer,
    SaveOCCIdentificationSerializer,
    SaveOCCLocationSerializer,
    SaveOCCObservationDetailSerializer,
    SaveOCCPlantCountSerializer,
    SaveOccurrenceDocumentSerializer,
    SaveOccurrenceReportDocumentSerializer,
    SaveOccurrenceReportSerializer,
    SaveOccurrenceSerializer,
    SaveOccurrenceSiteSerializer,
    SaveOCCVegetationStructureSerializer,
    SaveOCRAnimalObservationSerializer,
    SaveOCRAssociatedSpeciesSerializer,
    SaveOCRConservationThreatSerializer,
    SaveOCRFireHistorySerializer,
    SaveOCRHabitatCompositionSerializer,
    SaveOCRHabitatConditionSerializer,
    SaveOCRIdentificationSerializer,
    SaveOCRLocationSerializer,
    SaveOCRObservationDetailSerializer,
    SaveOCRPlantCountSerializer,
    SaveOCRVegetationStructureSerializer,
    SiteGeometrySerializer,
)
from boranga.components.occurrence.utils import (
    get_all_related_species,
    ocr_proposal_submit,
    process_shapefile_document,
    validate_map_files,
)
from boranga.components.spatial.utils import (
    populate_occurrence_tenure_data,
    save_geometry,
    spatially_process_geometry,
    transform_json_geometry,
)
from boranga.components.species_and_communities.models import GroupType, Taxonomy
from boranga.components.species_and_communities.serializers import TaxonomySerializer
from boranga.helpers import (
    is_contributor,
    is_customer,
    is_external_contributor,
    is_internal,
    is_internal_contributor,
    is_occurrence_approver,
    is_occurrence_assessor,
    is_occurrence_report_referee,
    is_readonly_user,
)

logger = logging.getLogger(__name__)


class OccurrenceReportFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        if view.name and "internal" in view.name:
            total_count = queryset.count()

            filter_group_type = request.GET.get("filter_group_type")
            if filter_group_type and not filter_group_type.lower() == "all":
                queryset = queryset.filter(group_type__name=filter_group_type)

            filter_occurrence = request.GET.get("filter_occurrence")
            if filter_occurrence and not filter_occurrence.lower() == "all":
                queryset = queryset.filter(occurrence_id=filter_occurrence)

            filter_scientific_name = request.GET.get("filter_scientific_name")
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(species__taxonomy__id=filter_scientific_name)

            filter_community_name = request.GET.get("filter_community_name")
            if filter_community_name and not filter_community_name.lower() == "all":
                queryset = queryset.filter(
                    community__taxonomy__id=filter_community_name
                )

            filter_status = request.GET.get("filter_status")
            if filter_status and not filter_status.lower() == "all":
                queryset = queryset.filter(processing_status=filter_status)

            def get_date(filter_date):
                date = request.GET.get(filter_date)
                if date:
                    date = datetime.strptime(date, "%Y-%m-%d")
                return date

            filter_submitted_from_date = get_date("filter_submitted_from_date")
            filter_submitted_to_date = get_date("filter_submitted_to_date")
            if filter_submitted_to_date:
                filter_submitted_to_date = datetime.combine(
                    filter_submitted_to_date, time.max
                )

            if filter_submitted_from_date and not filter_submitted_to_date:
                queryset = queryset.filter(
                    lodgement_date__gte=filter_submitted_from_date
                )

            if filter_submitted_from_date and filter_submitted_to_date:
                queryset = queryset.filter(
                    lodgement_date__range=[
                        filter_submitted_from_date,
                        filter_submitted_to_date,
                    ]
                )

            if filter_submitted_to_date and not filter_submitted_from_date:
                queryset = queryset.filter(lodgement_date__lte=filter_submitted_to_date)

            filter_from_observation_date = request.GET.get(
                "filter_observation_from_date"
            )
            filter_to_observation_date = request.GET.get("filter_observation_to_date")

            if filter_from_observation_date:
                queryset = queryset.filter(
                    observation_date__gte=filter_from_observation_date
                )
            if filter_to_observation_date:
                queryset = queryset.filter(
                    observation_date__lte=filter_to_observation_date
                )

        else:
            total_count = queryset.count()

            filter_group_type = request.GET.get("filter_group_type")
            if filter_group_type and not filter_group_type.lower() == "all":
                queryset = queryset.filter(group_type__name=filter_group_type)

            # filter_scientific_name is the species_id
            filter_scientific_name = request.GET.get("filter_scientific_name")
            if filter_scientific_name and not filter_scientific_name.lower() == "all":
                queryset = queryset.filter(species=filter_scientific_name)

            # filter_community_name is the community_id
            filter_community_name = request.GET.get("filter_community_name")
            if filter_community_name and not filter_community_name.lower() == "all":
                queryset = queryset.filter(community=filter_community_name)

            filter_application_status = request.GET.get("filter_application_status")
            if (
                filter_application_status
                and not filter_application_status.lower() == "all"
            ):
                queryset = queryset.filter(customer_status=filter_application_status)

        fields = self.get_fields(request)

        search_text = request.GET.get("search[value]")
        search_queryset = None

        # for search values that cannot be accommodated by DRF
        if search_text and "internal" in view.name:
            observer_ids = (
                OCRObserverDetail.objects.filter(main_observer=True)
                .filter(observer_name__icontains=search_text)
                .values_list("occurrence_report__id", flat=True)
            )
            search_queryset = queryset.filter(
                Q(submitter_information__name__icontains=search_text)
                | Q(id__in=observer_ids)
            )

        queryset = super().filter_queryset(request, queryset, view)

        if search_queryset:
            queryset = search_queryset.union(queryset)

        ordering = self.get_ordering(request, view, fields)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OccurrenceReportPaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    filter_backends = (OccurrenceReportFilterBackend,)
    pagination_class = DatatablesPageNumberPagination
    queryset = OccurrenceReport.objects.none()
    serializer_class = ListOccurrenceReportSerializer
    page_size = 10

    def get_serializer_class(self):
        if self.action == "occurrence_report_internal":
            return ListInternalOccurrenceReportSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = super().get_queryset()
        if is_internal(self.request):
            qs = OccurrenceReport.objects.all()
        elif is_contributor(self.request):
            qs = OccurrenceReport.objects.filter(submitter=self.request.user.id)

        return qs

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_external(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = qs.filter(internal_application=False)
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListOccurrenceReportSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListInternalOccurrenceReportSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_external_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "group_type",
            "scientific_name",
            "community_name",
            "customer_status",
            "occurrence_report_number",
        ]

        serializer = ListOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Type",
            "Scientific Name",
            "Community Name",
            "Status",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Type",
            "Scientific Name",
            "Community Name",
            "Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ExternalOccurrenceReports.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_ExternalOccurrenceReports.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_report_internal_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "scientific_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
            "occurrence_name",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
            "Occurrence",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def community_occurrence_report_internal_export(self, request, *args, **kwargs):

        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "community_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
            "occurrence_name",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Community Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
            "Occurrence",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Community Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Community.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Community.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def referred_to_me(self, request, *args, **kwargs):
        self.serializer_class = DTOccurrenceReportReferralSerializer
        qs = OccurrenceReportReferral.objects.filter(referral=request.user.id)
        self.filter_backends = (OccurrenceReportReferralFilterBackend,)
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = DTOccurrenceReportReferralSerializer(
            result_page, context={"request": request}, many=True
        )
        return self.paginator.get_paginated_response(serializer.data)


class OccurrenceReportViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin, DatumSearchMixin
):
    queryset = OccurrenceReport.objects.none()
    serializer_class = OccurrenceReportSerializer
    lookup_field = "id"

    def get_queryset(self):
        request = self.request
        qs = self.queryset
        if not is_internal(request) and not is_contributor(request):
            return qs

        if is_internal(request):
            qs = OccurrenceReport.objects.all()
        elif is_contributor(request) and is_occurrence_report_referee(request):
            qs = OccurrenceReport.objects.filter(
                Q(submitter=request.user.id) | Q(referrals__referral=request.user.id)
            )
        elif is_contributor(request):
            qs = OccurrenceReport.objects.filter(submitter=request.user.id)
        elif is_occurrence_report_referee(request):
            qs = OccurrenceReport.objects.filter(referrals__referral=request.user.id)

        return qs

    def get_serializer_class(self):
        if is_internal(self.request):
            return InternalOccurrenceReportSerializer
        return OccurrenceReportSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        group_type_id = GroupType.objects.get(id=request.data.get("group_type_id"))

        new_instance = OccurrenceReport(
            submitter=request.user.id,
            group_type=group_type_id,
        )
        if is_internal(request):
            new_instance.internal_application = True

        new_instance.save(version_user=request.user)
        data = {"occurrence_report_id": new_instance.id}

        # create Location for new instance
        serializer = SaveOCRLocationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatComposition for new instance
        serializer = SaveOCRHabitatCompositionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatCondition for new instance
        serializer = SaveOCRHabitatConditionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCRFireHistorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCRAssociatedSpeciesSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create ObservationDetail for new instance
        serializer = SaveOCRObservationDetailSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create PlantCount for new instance
        serializer = SaveOCRPlantCountSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AnimalObservation for new instance
        serializer = SaveOCRAnimalObservationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create Identification for new instance
        serializer = SaveOCRIdentificationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        serialized_obj = CreateOccurrenceReportSerializer(new_instance)
        return Response(serialized_obj.data)

    @detail_route(
        methods=[
            "PATCH",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.discard(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "PATCH",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.reinstate(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="transform-geometry",
    )
    def transform_geometry(self, request, *args, **kwargs):
        geometry = request.GET.get("geometry", None)
        from_srid = int(request.GET.get("from", 4326))
        to_srid = int(request.GET.get("to", 4326))

        if not geometry:
            return HttpResponse({}, content_type="application/json")

        json_geom = json.loads(geometry)

        transformed = transform_json_geometry(json_geom, from_srid, to_srid)

        return HttpResponse(transformed, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="spatially-process-geometries",
    )
    def spatially_process_geometries(self, request, *args, **kwargs):
        geometry = request.GET.get("geometry", None)
        operation = request.GET.get("operation", None)
        parameters = request.GET.get("parameters", None)
        parameters = [float(p) for p in parameters.split(",")] if parameters else []
        unit = request.GET.get("unit", None)

        if not geometry:
            raise serializers.ValidationError("Geometry is required")
        if not operation:
            raise serializers.ValidationError("Operation is required")
        if not unit:
            raise serializers.ValidationError("Unit is required")

        try:
            res_json = spatially_process_geometry(
                json.loads(geometry), operation, parameters, unit
            )
        except Exception as e:
            raise e
        else:
            return HttpResponse(res_json, content_type="application/json")

    # used for Location Tab of Occurrence Report external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="location-list-of-values",
    )
    def location_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        qs = self.get_queryset()
        datum_list = []

        id = request.GET.get("id", None)
        try:
            qs = qs.get(id=id)
        except OccurrenceReport.DoesNotExist:
            logger.error(f"Occurrence Report with id {id} not found")
        else:
            ocr_geometries = qs.ocr_geometry.all().exclude(**{"geometry": None})
            epsg_codes = [
                str(g.srid)
                for g in ocr_geometries.values_list("geometry", flat=True).distinct()
            ]
            # Add the srids of the original geometries to epsg_codes
            original_geometry_srids = [
                str(g.original_geometry_srid) for g in ocr_geometries
            ]
            epsg_codes += [g for g in original_geometry_srids if g.isnumeric()]
            epsg_codes = list(set(epsg_codes))
            datum_list = search_datums("", codes=epsg_codes)

        coordinate_source_list = []
        values = CoordinateSource.objects.all()
        if values:
            for val in values:
                coordinate_source_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        location_accuracy_list = []
        values = LocationAccuracy.objects.all()
        if values:
            for val in values:
                location_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "datum_list": datum_list,
            "coordinate_source_list": coordinate_source_list,
            "location_accuracy_list": location_accuracy_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    # used for Occurrence Report external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        land_form_list = []
        types = LandForm.objects.all()
        if types:
            for val in types:
                land_form_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        rock_type_list = []
        types = RockType.objects.all()
        if types:
            for val in types:
                rock_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_type_list = []
        types = SoilType.objects.all()
        if types:
            for val in types:
                soil_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_colour_list = []
        colours = SoilColour.objects.all()
        if colours:
            for val in colours:
                soil_colour_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_condition_list = []
        conditions = SoilCondition.objects.all()
        if conditions:
            for val in conditions:
                soil_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        drainage_list = []
        drainages = Drainage.objects.all()
        if drainages:
            for val in drainages:
                drainage_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        intensity_list = []
        intensities = Intensity.objects.all()
        if intensities:
            for val in intensities:
                intensity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "land_form_list": land_form_list,
            "rock_type_list": rock_type_list,
            "soil_type_list": soil_type_list,
            "soil_colour_list": soil_colour_list,
            "soil_condition_list": soil_condition_list,
            "drainage_list": drainage_list,
            "intensity_list": intensity_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def section_values(self, request, *args, **kwargs):

        section = request.GET.get("section")
        ocr = self.get_object()
        res_json = {}

        if hasattr(ocr, section):
            section_value = getattr(ocr, section)
            section_fields = section_value._meta.get_fields()

            for i in section_fields:
                if (
                    i.name == "id"
                    or i.name == "occurrence_report"
                    or isinstance(i, models.ManyToOneRel)
                ):
                    continue

                # ensure many to many fields are assigned an appropriate __str__
                if isinstance(i, models.ManyToManyField):
                    sub_section_values = getattr(section_value, i.name)
                    res_json[i.name] = []
                    for j in sub_section_values.all():
                        if j.__str__():
                            res_json[i.name].append(j.__str__())
                        else:
                            res_json[i.name].append(j.id)

                elif isinstance(i, models.ForeignKey):
                    sub_section_value = getattr(section_value, i.name)
                    if sub_section_value is not None:
                        res_json[i.name] = {}
                        sub_section_fields = sub_section_value._meta.get_fields()
                        for j in sub_section_fields:
                            if (
                                j.name != "id"
                                and not isinstance(j, models.ForeignKey)
                                and not isinstance(j, models.ManyToOneRel)
                                and not isinstance(j, models.ManyToManyRel)
                                and getattr(sub_section_value, j.name) is not None
                            ):
                                res_json[i.name][j.name] = str(
                                    getattr(sub_section_value, j.name)
                                )
                        # if the num sub section has only one value, assign as section
                        if len(res_json[i.name]) == 1:
                            res_json[i.name] = list(res_json[i.name].values())[0]
                elif isinstance(i, MultiSelectField):
                    if i.choices:
                        choice_dict = dict(i.choices)
                        id_list = getattr(section_value, i.name)
                        values_list = []
                        for id in id_list:
                            if id.isdigit() and int(id) in choice_dict:
                                values_list.append(choice_dict[int(id)])
                        res_json[i.name] = values_list
                    else:
                        res_json[i.name] = getattr(section_value, i.name)

                elif getattr(section_value, i.name) is not None:
                    res_json[i.name] = str(getattr(section_value, i.name))

        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @detail_route(methods=["get"], detail=True)
    def add_related_species(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        if instance.associated_species:
            related_species = instance.associated_species.related_species
        else:
            raise serializers.ValidationError("Associated Species does not exist")

        taxon_id = request.GET.get("species")

        try:
            taxon = Taxonomy.objects.get(id=taxon_id)
        except Taxonomy.DoesNotExist:
            raise serializers.ValidationError("Species does not exist")

        if taxon not in related_species.all():
            related_species.add(taxon)
        else:
            raise serializers.ValidationError("Species already added")

        instance.save(version_user=request.user)

        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )

        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def remove_related_species(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        if instance.associated_species:
            related_species = instance.associated_species.related_species
        else:
            raise serializers.ValidationError("Associated Species does not exist")

        taxon_id = request.GET.get("species")

        try:
            taxon = Taxonomy.objects.get(id=taxon_id)
        except Taxonomy.DoesNotExist:
            raise serializers.ValidationError("Species does not exist")

        if taxon in related_species.all():
            related_species.remove(taxon)
        else:
            raise serializers.ValidationError("Species not related")

        instance.save(version_user=request.user)

        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )

        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_species(self, request, *args, **kwargs):
        instance = self.get_object()
        if hasattr(instance, "associated_species"):
            related_species = instance.associated_species.related_species
        else:
            related_species = Taxonomy.objects.none()
        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )
        return Response(serializer.data)

    # used for Occurrence Report Observation external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def observation_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        observation_method_list = []
        values = ObservationMethod.objects.all()
        if values:
            for val in values:
                observation_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_method_list = []
        values = PlantCountMethod.objects.all()
        if values:
            for val in values:
                plant_count_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_accuracy_list = []
        values = PlantCountAccuracy.objects.all()
        if values:
            for val in values:
                plant_count_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_condition_list = []
        values = PlantCondition.objects.all()
        if values:
            for val in values:
                plant_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        counted_subject_list = []
        values = CountedSubject.objects.all()
        if values:
            for val in values:
                counted_subject_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        primary_detection_method_list = []
        values = PrimaryDetectionMethod.objects.all()
        if values:
            for val in values:
                primary_detection_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        secondary_sign_list = []
        values = SecondarySign.objects.all()
        if values:
            for val in values:
                secondary_sign_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        reprod_state_list = []
        values = ReproductiveState.objects.all()
        if values:
            for val in values:
                reprod_state_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        death_reason_list = []
        values = DeathReason.objects.all()
        if values:
            for val in values:
                death_reason_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        animal_health_list = []
        values = AnimalHealth.objects.all()
        if values:
            for val in values:
                animal_health_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        identification_certainty_list = []
        values = IdentificationCertainty.objects.all()
        if values:
            for val in values:
                identification_certainty_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_type_list = []
        values = SampleType.objects.all()
        if values:
            for val in values:
                sample_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_dest_list = []
        values = SampleDestination.objects.all()
        if values:
            for val in values:
                sample_dest_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        permit_type_list = []
        values = PermitType.objects.all()
        if values:
            for val in values:
                permit_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "observation_method_list": observation_method_list,
            "plant_count_method_list": plant_count_method_list,
            "plant_count_accuracy_list": plant_count_accuracy_list,
            "plant_condition_list": plant_condition_list,
            "counted_subject_list": counted_subject_list,
            "primary_detection_method_list": primary_detection_method_list,
            "secondary_sign_list": secondary_sign_list,
            "reprod_state_list": reprod_state_list,
            "death_reason_list": death_reason_list,
            "animal_health_list": animal_health_list,
            "identification_certainty_list": identification_certainty_list,
            "sample_type_list": sample_type_list,
            "sample_dest_list": sample_dest_list,
            "permit_type_list": permit_type_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    def is_authorised_to_update(self):
        # To update an occurrence report, the user must be:
        # - the original submitter and the OCR in draft or
        # - an internal assessor and the OCR under assessment or
        instance = self.get_object()
        user = self.request.user
        if not (
            (
                instance.can_user_edit
                and (
                    user.id
                    == instance.submitter  # or
                    # (instance.internal_application and is_internal(self.request))
                )
            )
            or (instance.has_assessor_mode(self.request))
            or (instance.has_unlocked_mode(self.request))
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )

    def is_authorised_to_assign(self, assigner, assignee=None):
        # To assign a report:
        # - the report must be under assessment, the assigner must be in the assessment group,
        # and the assignee must be in the assessment group or
        # - the report must be under approval, the assigner must be in the approver group,
        # and the assignee must be in the approval group
        # AND the Assignee must be the proposed assignee, or already assigned
        instance = self.get_object()

        in_assessor_group = assignee and is_occurrence_assessor(self.request)
        in_approver_group = assignee and is_occurrence_approver(self.request)

        self_assigning = assigner == assignee

        assigner_assigned = instance.assigned_officer == assigner.id
        assigner_approver = instance.assigned_approver == assigner.id

        if (
            instance.processing_status
            in [
                OccurrenceReport.PROCESSING_STATUS_WITH_REFERRAL,
                OccurrenceReport.PROCESSING_STATUS_WITH_ASSESSOR,
                OccurrenceReport.PROCESSING_STATUS_UNLOCKED,
            ]
        ) and (
            (self_assigning and (in_assessor_group or in_approver_group))
            or (
                not (assignee)
                and assigner_assigned
                and instance.has_assessor_mode(self.request)
            )
            or (
                (in_assessor_group or in_approver_group)
                and assigner_assigned
                and instance.has_assessor_mode(self.request)
            )
        ):
            return
        elif (
            instance.processing_status
            in [
                OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER,
            ]
        ) and (
            (self_assigning and in_approver_group)
            or (
                not (assignee)
                and assigner_approver
                and instance.has_approver_mode(self.request)
            )
            or (
                (in_approver_group)
                and assigner_assigned
                and instance.has_assessor_mode(self.request)
            )
        ):
            return

        raise serializers.ValidationError(
            "User not authorised to manage assignments for Occurrence Report"
        )

    def is_authorised_to_assess(self):
        instance = self.get_object()
        if not instance.has_assessor_mode(self.request):
            raise serializers.ValidationError(
                "User not authorised to make Assessment Actions for Occurrence Report"
            )

    def is_authorised_to_approve(self):
        instance = self.get_object()
        if not instance.has_approver_mode(self.request):
            raise serializers.ValidationError(
                "User not authorised to make Approval Actions for Occurrence Report"
            )

    def is_authorised_to_change_lock(self):
        instance = self.get_object()

        if not instance.can_change_lock(self.request):
            raise serializers.ValidationError(
                "User not authorised to change lock status for Occurrence Report"
            )

    def unlocked_back_to_assessor(self):
        instance = self.get_object()
        request = self.request
        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            serializer = BackToAssessorSerializer(
                data={"reason": "Change made after unlock"}
            )
            serializer.is_valid(raise_exception=True)
            instance.back_to_assessor(request, serializer.validated_data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def lock_occurrence_report(self, request, *args, **kwargs):
        self.is_authorised_to_change_lock()
        instance = self.get_object()
        instance.lock(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def unlock_occurrence_report(self, request, *args, **kwargs):
        self.is_authorised_to_change_lock()
        instance = self.get_object()
        instance.unlock(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_location_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()

        location_instance, created = OCRLocation.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # species_id saved seperately as its not field of Location but OCR
        # species = request.data.get("species_id")
        # ocr_instance.species_id = species
        # ocr_instance.save()
        # community_id saved seperately as its not field of Location but OCR
        # community = request.data.get("community_id")
        # ocr_instance.community_id = community

        # if ocr_instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
        #    self.unlocked_back_to_assessor()
        # else:
        #    ocr_instance.save(version_user=request.user)

        # ocr geometry data to save seperately
        geometry_data = request.data.get("ocr_geometry")
        if geometry_data:
            save_geometry(request, ocr_instance, geometry_data, "occurrence_report")

        # polygon = request.data.get('geojson_polygon')
        # if polygon:
        #     coords_list = [list(map(float, coord.split(' '))) for coord in polygon.split(',')]
        #     coords_list.append(coords_list[0])
        #     request.data['geojson_polygon'] = GEOSGeometry(f'POLYGON(({", ".join(map(lambda
        # x: " ".join(map(str, x)), coords_list))}))')

        # the request.data is only the habitat composition data thats been sent from front end
        location_data = request.data.get("location")
        serializer = SaveOCRLocationSerializer(
            location_instance, data=location_data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_composition_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()

        habitat_instance, created = OCRHabitatComposition.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRHabitatCompositionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_condition_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        habitat_instance, created = OCRHabitatCondition.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCRHabitatConditionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_vegetation_structure(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        vegetation_instance, created = OCRVegetationStructure.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCRVegetationStructureSerializer(
            vegetation_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_fire_history_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        fire_instance, created = OCRFireHistory.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRFireHistorySerializer(
            fire_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_associated_species_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        assoc_species_instance, created = OCRAssociatedSpecies.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCRAssociatedSpeciesSerializer(
            assoc_species_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_observation_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        obs_det_instance, created = OCRObservationDetail.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the observation detail data thats been sent from front end
        serializer = SaveOCRObservationDetailSerializer(
            obs_det_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_plant_count_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        plant_count_instance, created = OCRPlantCount.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the plant count data thats been sent from front end
        serializer = SaveOCRPlantCountSerializer(
            plant_count_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_animal_observation_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        animal_obs_instance, created = OCRAnimalObservation.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the animal obs data thats been sent from front end
        serializer = SaveOCRAnimalObservationSerializer(
            animal_obs_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_identification_details(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        ocr_instance = self.get_object()
        identification_instance, created = OCRIdentification.objects.get_or_create(
            occurrence_report=ocr_instance
        )
        # the request.data is only the identification data thats been sent from front end
        serializer = SaveOCRIdentificationSerializer(
            identification_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        if (
            ocr_instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor()

        return Response(serializer.data)

    # used for observer detail datatable
    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def observer_details(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.observer_detail.all()
        serializer = OCRObserverDetailLimitedSerializer(
            qs, many=True, context={"request": request}
        )
        if (
            is_occurrence_assessor(request)
            or is_occurrence_approver(request)
            or is_occurrence_report_referee(request, instance)
            or ((is_contributor(request)) and instance.submitter == request.user.id)
        ):
            serializer = OCRObserverDetailSerializer(
                qs, many=True, context={"request": request}
            )

        return Response(serializer.data)

    @list_route(methods=["GET"], detail=False)
    def list_for_map(self, request, *args, **kwargs):
        """Returns the proposals for the map"""
        occurrence_report_ids = [
            int(id)
            for id in request.query_params.get("proposal_ids", "").split(",")
            if id.lstrip("-").isnumeric()
        ]
        # application_type = request.query_params.get("application_type", None)
        # processing_status = request.query_params.get("processing_status", None)

        cache_key = settings.CACHE_KEY_MAP_OCCURRENCE_REPORTS
        qs = cache.get(cache_key)

        if qs is None:
            qs = (
                self.get_queryset()
                .exclude(ocr_geometry__isnull=True)
                .prefetch_related("ocr_geometry")
            )
            cache.set(cache_key, qs, settings.CACHE_TIMEOUT_2_HOURS)

        if len(occurrence_report_ids) > 0:
            qs = qs.filter(id__in=occurrence_report_ids)

        # if (
        #     application_type
        #     and application_type.isnumeric()
        #     and int(application_type) > 0
        # ):
        #     qs = qs.filter(application_type_id=application_type)

        # if processing_status:
        #     qs = qs.filter(processing_status=processing_status)

        # qs = self.filter_queryset(qs)
        serializer = ListOCRReportMinimalSerializer(
            qs, context={"request": request}, many=True
        )
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def draft(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        instance = self.get_object()
        # request_data = request.data
        proposal_data = (
            request.data.get("proposal") if request.data.get("proposal") else {}
        )
        # request.data['submitter'] = u'{}'.format(request.user.id)
        if "submitter" in proposal_data and proposal_data["submitter"]:
            request.data.get("proposal")["submitter"] = "{}".format(
                proposal_data["submitter"].get("id")
            )
        if proposal_data.get("habitat_composition"):
            habitat_instance, created = OCRHabitatComposition.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRHabitatCompositionSerializer(
                habitat_instance, data=proposal_data.get("habitat_composition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("habitat_condition"):
            hab_cond_instance, created = OCRHabitatCondition.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRHabitatConditionSerializer(
                hab_cond_instance, data=proposal_data.get("habitat_condition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("vegetation_structure"):
            veg_struct_instance, created = OCRVegetationStructure.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRVegetationStructureSerializer(
                veg_struct_instance, data=proposal_data.get("vegetation_structure")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("fire_history"):
            fire_instance, created = OCRFireHistory.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRFireHistorySerializer(
                fire_instance, data=proposal_data.get("fire_history")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("associated_species"):
            assoc_species_instance, created = (
                OCRAssociatedSpecies.objects.get_or_create(occurrence_report=instance)
            )
            serializer = SaveOCRAssociatedSpeciesSerializer(
                assoc_species_instance,
                data=proposal_data.get("associated_species"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("observation_detail"):
            obs_det_instance, created = OCRObservationDetail.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRObservationDetailSerializer(
                obs_det_instance, data=proposal_data.get("observation_detail")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("plant_count"):
            plant_count_instance, created = OCRPlantCount.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRPlantCountSerializer(
                plant_count_instance, data=proposal_data.get("plant_count")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("animal_observation"):
            animal_obs_instance, created = OCRAnimalObservation.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRAnimalObservationSerializer(
                animal_obs_instance,
                data=proposal_data.get("animal_observation"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("identification"):
            identification_instance, created = OCRIdentification.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRIdentificationSerializer(
                identification_instance,
                data=proposal_data.get("identification"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if proposal_data.get("location"):
            location_instance, created = OCRLocation.objects.get_or_create(
                occurrence_report=instance
            )
            serializer = SaveOCRLocationSerializer(
                location_instance, data=proposal_data.get("location")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        # ocr geometry data to save seperately
        geometry_data = proposal_data.get("ocr_geometry", None)
        if geometry_data:
            save_geometry(request, instance, geometry_data, "occurrence_report")

        serializer = SaveOccurrenceReportSerializer(
            instance, data=proposal_data, partial=True
        )

        serializer.is_valid(raise_exception=True)
        if serializer.is_valid():
            if (
                instance.processing_status
                == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
            ):
                serializer.save(no_revision=True)
                self.unlocked_back_to_assessor()
            else:
                serializer.save(version_user=request.user)

        final_instance = self.get_object()
        serializer = self.get_serializer(final_instance)
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    def submit(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        instance = self.get_object()
        # instance.submit(request,self)
        ocr_proposal_submit(instance, request)
        instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
        # return redirect(reverse('external'))

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def action_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.action_logs.all()
        serializer = OccurrenceReportUserActionSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.comms_logs.all()
        serializer = OccurrenceReportLogEntrySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def add_comms_log(self, request, *args, **kwargs):

        if not is_internal(self):
            raise serializers.ValidationError(
                "User not authorised to add Communication Logs to the Occurrence Report"
            )

        instance = self.get_object()
        mutable = request.data._mutable
        request.data._mutable = True
        request.data["occurrence_report"] = f"{instance.id}"
        request.data["staff"] = f"{request.user.id}"
        request.data._mutable = mutable
        serializer = OccurrenceReportLogEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comms = serializer.save()
        # Save the files
        for f in request.FILES:
            document = comms.documents.create()
            document.name = str(request.FILES[f])
            document._file = request.FILES[f]
            document.save()
        # End Save Documents

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()

        if not is_internal and not is_external_contributor(request):
            raise PermissionDenied  # TODO: Replace with permission class

        qs = instance.documents.all()
        qs = qs.exclude(input_name="occurrence_report_approval_doc")
        if not is_internal(request) and is_external_contributor(request):
            qs = qs.filter(
                occurrence_report__submitter=self.request.user.id,
                visible=True,
                can_submitter_access=True,
            )

        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceReportDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.ocr_threats.all()
        if is_internal(self.request):
            qs = instance.ocr_threats.all()
        elif is_customer(self.request):
            # TODO Do we need to sort the threats for external user (similar like documents)
            # qs = qs.filter(Q(uploaded_by=request.user.id))
            qs = instance.ocr_threats.all()
        filter_backend = OCCConservationThreatFilterBackend()
        qs = filter_backend.filter_queryset(self.request, qs, self)
        serializer = OCRConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["POST"], detail=True)
    @renderer_classes((JSONRenderer,))
    def process_shapefile_document(self, request, *args, **kwargs):
        instance = self.get_object()
        returned_data = None
        returned_data = process_shapefile_document(request, instance)
        if returned_data:
            return Response(returned_data)
        else:
            return Response({})

    @detail_route(methods=["POST"], detail=True)
    @renderer_classes((JSONRenderer,))
    def validate_map_files(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        validate_map_files(request, instance, "occurrence_report")
        if instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.unlocked_back_to_assessor()
            instance.save(no_revision=True)
        else:
            instance.save(version_user=request.user)
        serializer = self.get_serializer(instance)

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def assign_request_user(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_assign(request.user, request.user)
        instance.assign_officer(request, request.user)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def assign_to(self, request, *args, **kwargs):

        instance = self.get_object()
        user_id = request.data.get("assessor_id", None)
        user = None
        if not user_id:
            raise serializers.ValidationError("An assessor id is required")
        try:
            user = EmailUser.objects.get(id=user_id)
        except EmailUser.DoesNotExist:
            raise serializers.ValidationError(
                "A user with the id passed in does not exist"
            )
        assigner = self.request.user
        self.is_authorised_to_assign(assigner, user)
        instance.assign_officer(request, user)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def unassign(self, request, *args, **kwargs):

        user = self.request.user
        self.is_authorised_to_assign(user)

        instance = self.get_object()
        instance.unassign(request)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def amendment_request(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.amendment_requests
        qs = qs.filter(status="requested")
        serializer = OccurrenceReportAmendmentRequestSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def propose_decline(self, request, *args, **kwargs):

        self.is_authorised_to_assess()

        instance = self.get_object()
        serializer = ProposeDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.propose_decline(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def decline(self, request, *args, **kwargs):

        self.is_authorised_to_approve()

        instance = self.get_object()

        original_occ = instance.occurrence

        serializer = ProposeDeclineSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.decline(request, serializer.validated_data)  # ensure occ set to None

        # run occ check
        if original_occ:
            original_occ.check_ocr_count_for_discard(request)

        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def back_to_assessor(self, request, *args, **kwargs):
        instance = self.get_object()

        if (
            instance.processing_status
            == OccurrenceReport.PROCESSING_STATUS_WITH_APPROVER
        ):
            self.is_authorised_to_approve()
        elif instance.processing_status == OccurrenceReport.PROCESSING_STATUS_UNLOCKED:
            self.is_authorised_to_update()

        serializer = BackToAssessorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.back_to_assessor(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def propose_approve(self, request, *args, **kwargs):

        self.is_authorised_to_assess()

        instance = self.get_object()
        serializer = ProposeApproveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance.propose_approve(request, serializer.validated_data)
        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def approve(self, request, *args, **kwargs):

        self.is_authorised_to_approve()

        instance = self.get_object()

        original_occ = instance.occurrence

        instance.approve(request)

        if original_occ and original_occ.id != instance.occurrence.id:
            original_occ.check_ocr_count_for_discard(request)

        serializer = InternalOccurrenceReportSerializer(
            instance, context={"request": request}
        )
        return Response(serializer.data)

    # used on referral form
    @detail_route(methods=["post"], detail=True)
    def send_referral(self, request, *args, **kwargs):
        self.is_authorised_to_assess()
        instance = self.get_object()
        serializer = SendReferralSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        instance.send_referral(
            request,
            serializer.validated_data["email"],
            serializer.validated_data["text"],
        )
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def referrals(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.referrals.all()
        serializer = OccurrenceReportProposalReferralSerializer(
            qs, many=True, context={"request": self.request}
        )
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    def external_referee_invite(self, request, *args, **kwargs):
        instance = self.get_object()
        request.data["occurrence_report_id"] = instance.id
        serializer = OCRExternalRefereeInviteSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        if OCRExternalRefereeInvite.objects.filter(
            archived=False, email=request.data["email"]
        ).exists():
            raise serializers.ValidationError(
                "An external referee invitation has already been sent to {email}".format(
                    email=request.data["email"]
                ),
                code="invalid",
            )
        external_referee_invite = OCRExternalRefereeInvite.objects.create(
            sent_by=request.user.id, **request.data
        )
        send_external_referee_invite_email(instance, request, external_referee_invite)

        serializer_class = self.get_serializer_class()
        serializer = serializer_class(instance, context={"request": request})
        return Response(serializer.data)


class ObserverDetailViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OCRObserverDetail.objects.none()
    serializer_class = OCRObserverDetailLimitedSerializer

    def get_serializer_class(self):
        if (
            is_occurrence_assessor(self.request)
            or is_occurrence_approver(self.request)
            or is_external_contributor(self.request)
            or is_internal_contributor(self.request)
            or is_readonly_user(self.request)
        ):
            return OCRObserverDetailSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = OCRObserverDetail.objects.none()

        if (
            is_occurrence_assessor(self.request)
            or is_occurrence_approver(self.request)
            or is_readonly_user(self.request)
        ):
            qs = OCRObserverDetail.objects.all().order_by("id")
        elif is_contributor(self.request) and is_occurrence_report_referee(
            self.request
        ):
            qs = OCRObserverDetail.objects.filter(
                Q(occurrence_report__submitter=self.request.user.id)
                | Q(occurrence_report__referrals__referral=self.request.user.id)
            ).order_by("id")
        elif is_contributor(self.request):
            qs = OCRObserverDetail.objects.filter(
                occurrence_report__submitter=self.request.user.id
            ).order_by("id")
        elif is_occurrence_report_referee(self.request):
            qs = OCRObserverDetail.objects.filter(
                occurrence_report__referrals__referral=self.request.user.id
            ).order_by("id")

        return qs

    def is_authorised_to_update(self, occurrence_report):
        user = self.request.user
        if not (
            (
                occurrence_report.can_user_edit
                and (
                    user.id
                    == occurrence_report.submitter  # or
                    # (occurrence_report.internal_application and is_internal(self.request))
                )
            )
            or (occurrence_report.has_assessor_mode(self.request))
            or (occurrence_report.has_unlocked_mode(self.request))
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )

    def unlocked_back_to_assessor(self, occurrence_report):
        request = self.request
        if (
            occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            serializer = BackToAssessorSerializer(
                data={"reason": "Change made after unlock"}
            )
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)

        if not instance.visible:
            raise serializers.ValidationError("Discarded observer cannot be updated.")

        serializer = OCRObserverDetailSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)

        occurrence_report = serializer.validated_data["occurrence_report"]
        observer_name = serializer.validated_data["observer_name"]

        if (
            OCRObserverDetail.objects.exclude(id=instance.id)
            .filter(
                Q(observer_name=observer_name)
                & Q(occurrence_report=occurrence_report)
                & Q(visible=True)
            )
            .exists()
        ):
            raise serializers.ValidationError(
                "Observer with this name already exists for this occurrence report"
            )

        serializer.save()

        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        # instance.community.log_user_action(CommunityUserAction.ACTION_ADD_THREAT.format(instance.threat_number,instance.community.community_number),request)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = OCRObserverDetailSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)

        observer_name = serializer.validated_data["observer_name"]

        if OCRObserverDetail.objects.filter(
            Q(observer_name=observer_name)
            & Q(occurrence_report=occurrence_report)
            & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Observer with this name already exists for this occurrence report"
            )

        serializer.save()

        if (
            occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(occurrence_report)

        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save()

        serializer = self.get_serializer(instance)
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True

        if OCRObserverDetail.objects.filter(
            Q(observer_name=instance.observer_name)
            & Q(occurrence_report=instance.occurrence_report)
            & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Active observer with this name already exists for this occurrence report"
            )

        instance.save()

        serializer = self.get_serializer(instance)
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)


class OccurrenceReportAmendmentRequestViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = OccurrenceReportAmendmentRequest.objects.none()
    serializer_class = OccurrenceReportAmendmentRequestSerializer

    def get_queryset(self):
        if is_internal(self.request):  # user.is_authenticated():
            qs = OccurrenceReportAmendmentRequest.objects.all().order_by("id")
            return qs
        return OccurrenceReportAmendmentRequest.objects.none()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=json.loads(request.data.get("data")))
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        if not (occurrence_report.has_assessor_mode(self.request)):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )
        instance = serializer.save()
        instance.add_documents(request)
        instance.generate_amendment(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    def delete_document(self, request, *args, **kwargs):
        instance = self.get_object()

        occurrence_report = instance.occurrence_report
        if not (occurrence_report.has_assessor_mode(self.request)):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )

        OccurrenceReportAmendmentRequestDocument.objects.get(
            id=request.data.get("id")
        ).delete()
        return Response(
            [
                dict(id=i.id, name=i.name, _file=i._file.url)
                for i in instance.cs_amendment_request_documents.all()
            ]
        )


class OccurrenceReportDocumentViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = OccurrenceReportDocument.objects.none()
    serializer_class = OccurrenceReportDocumentSerializer

    def is_authorised_to_update(self, occurrence_report):
        user = self.request.user
        if not (
            (
                occurrence_report.can_user_edit
                and (
                    user.id
                    == occurrence_report.submitter  # or
                    # (occurrence_report.internal_application and is_internal(self.request))
                )
            )
            or (occurrence_report.has_assessor_mode(self.request))
            or (occurrence_report.has_unlocked_mode(self.request))
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )

    def unlocked_back_to_assessor(self, occurrence_report):
        request = self.request
        if (
            occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            serializer = BackToAssessorSerializer(
                data={"reason": "Change made after unlock"}
            )
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def get_queryset(self):
        if is_internal(self.request):
            return OccurrenceReportDocument.objects.all().order_by("id")
        if is_external_contributor(self.request):
            return OccurrenceReportDocument.objects.filter(
                occurrence_report__submitter=self.request.user.id,
                visible=True,
                can_submitter_access=True,
            )
        return OccurrenceReportDocument.objects.none()

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_DISCARD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_REINSTATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        data = json.loads(request.data.get("data"))
        serializer = SaveOccurrenceReportDocumentSerializer(instance, data=data)
        if is_internal(self.request):
            serializer = InternalSaveOccurrenceReportDocumentSerializer(
                instance, data=data
            )

        serializer.is_valid(raise_exception=True)
        serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_UPDATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = json.loads(request.data.get("data"))
        serializer = SaveOccurrenceReportDocumentSerializer(data=data)
        if is_internal(self.request):
            serializer = InternalSaveOccurrenceReportDocumentSerializer(data=data)

        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)
        instance = serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id

        if is_external_contributor(self.request):
            instance.can_submitter_access = True

        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_ADD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)
        return Response(serializer.data)


class OCRConservationThreatFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):

        total_count = queryset.count()

        filter_threat_category = request.GET.get("filter_threat_category")
        if filter_threat_category and not filter_threat_category.lower() == "all":
            queryset = queryset.filter(threat_category_id=filter_threat_category)

        filter_threat_current_impact = request.GET.get("filter_threat_current_impact")
        if (
            filter_threat_current_impact
            and not filter_threat_current_impact.lower() == "all"
        ):
            queryset = queryset.filter(current_impact=filter_threat_current_impact)

        filter_threat_potential_impact = request.GET.get(
            "filter_threat_potential_impact"
        )
        if (
            filter_threat_potential_impact
            and not filter_threat_potential_impact.lower() == "all"
        ):
            queryset = queryset.filter(potential_impact=filter_threat_potential_impact)

        filter_threat_status = request.GET.get("filter_threat_status")
        if filter_threat_status and not filter_threat_status.lower() == "all":
            if filter_threat_status == "active":
                queryset = queryset.filter(visible=True)
            elif filter_threat_status == "removed":
                queryset = queryset.filter(visible=False)

        def get_date(filter_date):
            date = request.GET.get(filter_date)
            if date:
                date = datetime.strptime(date, "%Y-%m-%d")
            return date

        filter_observed_from_date = get_date("filter_observed_from_date")
        if filter_observed_from_date:
            queryset = queryset.filter(date_observed__gte=filter_observed_from_date)

        filter_observed_to_date = get_date("filter_observed_to_date")
        if filter_observed_to_date:
            queryset = queryset.filter(date_observed__lte=filter_observed_to_date)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)

        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OCRConservationThreatViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OCRConservationThreat.objects.none()
    serializer_class = OCRConservationThreatSerializer

    def is_authorised_to_update(self, occurrence_report):
        user = self.request.user
        if not (
            (
                occurrence_report.can_user_edit
                and (
                    user.id
                    == occurrence_report.submitter  # or
                    # (occurrence_report.internal_application and is_internal(self.request))
                )
            )
            or (occurrence_report.has_assessor_mode(self.request))
            or (occurrence_report.has_unlocked_mode(self.request))
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence Report"
            )

    def unlocked_back_to_assessor(self, occurrence_report):
        request = self.request
        if (
            occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            serializer = BackToAssessorSerializer(
                data={"reason": "Change made after unlock"}
            )
            serializer.is_valid(raise_exception=True)
            occurrence_report.back_to_assessor(request, serializer.validated_data)

    def get_queryset(self):
        request_user = self.request.user
        qs = OCRConservationThreat.objects.none()

        if is_internal(self.request):
            qs = OCRConservationThreat.objects.all().order_by("id")
        elif is_customer(self.request):
            # TODO filter qs as per added_by - using the OCR submitter for now
            qs = OCRConservationThreat.objects.filter(
                occurrence_report__submitter=request_user.id
            ).order_by("id")
            return qs
        return qs

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_DISCARD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_REINSTATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)

        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence_report)
        serializer = SaveOCRConservationThreatSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        serializer.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_UPDATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)

        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOCRConservationThreatSerializer(
            data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        occurrence_report = serializer.validated_data["occurrence_report"]
        self.is_authorised_to_update(occurrence_report)
        instance = serializer.save(version_user=request.user)
        if instance.occurrence_report:
            instance.occurrence_report.log_user_action(
                OccurrenceReportUserAction.ACTION_ADD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence_report.occurrence_report_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)

        if (
            instance.occurrence_report.processing_status
            == OccurrenceReport.PROCESSING_STATUS_UNLOCKED
        ):
            self.unlocked_back_to_assessor(instance.occurrence_report)

        return Response(serializer.data)


class GetOCCProfileDict(views.APIView):
    def get(self, request, format=None):
        group_type = request.GET.get("group_type", "")
        logger.debug(
            "group_type: %s" % group_type
        )  # TODO: Unused variable here. Use or remove.
        wild_status_list = list(WildStatus.objects.all().values("id", "name"))
        occurrence_source_list = list(Occurrence.OCCURRENCE_SOURCE_CHOICES)

        res_json = {
            "wild_status_list": wild_status_list,
            "occurrence_source_list": occurrence_source_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")


class OccurrenceFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        filter_group_type = request.GET.get("filter_group_type")
        if filter_group_type and not filter_group_type.lower() == "all":
            queryset = queryset.filter(group_type__name=filter_group_type)

        filter_occurrence_name = request.GET.get("filter_occurrence_name")
        if filter_occurrence_name and not filter_occurrence_name.lower() == "all":
            queryset = queryset.filter(occurrence_name=filter_occurrence_name)

        filter_scientific_name = request.GET.get("filter_scientific_name")
        if filter_scientific_name and not filter_scientific_name.lower() == "all":
            queryset = queryset.filter(species__taxonomy__id=filter_scientific_name)

        filter_community_name = request.GET.get("filter_community_name")
        if filter_community_name and not filter_community_name.lower() == "all":
            queryset = queryset.filter(community__taxonomy__id=filter_community_name)

        filter_status = request.GET.get("filter_status")
        if filter_status and not filter_status.lower() == "all":
            queryset = queryset.filter(processing_status=filter_status)

        filter_from_review_due_date = request.GET.get("filter_from_review_due_date")
        filter_to_review_due_date = request.GET.get("filter_to_review_due_date")

        if filter_from_review_due_date:
            queryset = queryset.filter(review_due_date__gte=filter_from_review_due_date)
        if filter_to_review_due_date:
            queryset = queryset.filter(review_due_date__lte=filter_to_review_due_date)

        fields = self.get_fields(request)

        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)

        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OccurrencePaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    pagination_class = DatatablesPageNumberPagination
    queryset = Occurrence.objects.none()
    serializer_class = OccurrenceSerializer
    page_size = 10
    filter_backends = (OccurrenceFilterBackend,)

    def get_serializer_class(self):
        if self.action in ["list", "occurrence_internal", "occurrence_external"]:
            return ListOccurrenceSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = Occurrence.objects.all()
        if is_customer(self.request):
            return Occurrence.objects.none()
        return qs

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        serializer = ListOccurrenceSerializer(
            result_page, context={"request": request}, many=True
        )

        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_internal_export(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)
        export_format = request.GET.get("export_format")
        allowed_fields = [
            "species",
            "scientific_name",
            "reported_date",
            "submitter",
            "processing_status",
            "occurrence_report_number",
        ]

        serializer = ListInternalOccurrenceReportSerializer(
            qs, context={"request": request}, many=True
        )
        serialized_data = serializer.data

        filtered_data = []
        for obj in serialized_data:
            filtered_obj = {
                key: value for key, value in obj.items() if key in allowed_fields
            }
            filtered_data.append(filtered_obj)

        def flatten_dict(d, parent_key="", sep="_"):
            flattened_dict = {}
            for k, v in d.items():
                new_key = parent_key + sep + k if parent_key else k
                if isinstance(v, dict):
                    flattened_dict.update(flatten_dict(v, new_key, sep))
                else:
                    flattened_dict[new_key] = v
            return flattened_dict

        flattened_data = [flatten_dict(item) for item in filtered_data]
        df = pd.DataFrame(flattened_data)
        new_headings = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df.columns = new_headings
        column_order = [
            "Number",
            "Occurrence",
            "Scientific Name",
            "Submission date/time",
            "Submitter",
            "Processing Status",
        ]
        df = df[column_order]

        if export_format is not None:
            if export_format == "excel":
                buffer = BytesIO()
                workbook = Workbook()
                sheet_name = "Sheet1"
                sheet = workbook.active
                sheet.title = sheet_name

                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(buffer)
                buffer.seek(0)
                response = HttpResponse(
                    buffer.read(), content_type="application/vnd.ms-excel"
                )
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.xlsx"
                )
                final_response = response
                buffer.close()
                return final_response

            elif export_format == "csv":
                csv_data = df.to_csv(index=False)
                response = HttpResponse(content_type="text/csv")
                response["Content-Disposition"] = (
                    "attachment; filename=DBCA_OccurrenceReport_Species.csv"
                )
                response.write(csv_data)
                return response

            else:
                return Response(status=400, data="Format not valid")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        group_type_id = request.GET.get("group_type_id", None)
        if group_type_id:
            queryset = queryset.filter(group_type_id=group_type_id)
        search_term = request.GET.get("term", "")
        if search_term:
            queryset = queryset.values_list("occurrence_number", flat=True)
            queryset = (
                queryset.filter(occurrence_number__icontains=search_term)
                .distinct()
                .values("id", "occurrence_number")[:10]
            )
            queryset = [
                {"id": occurrence["id"], "text": occurrence["occurrence_number"]}
                for occurrence in queryset
            ]
        return Response({"results": queryset})

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_name_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset().filter(
            processing_status=Occurrence.PROCESSING_STATUS_ACTIVE
        )
        group_type_id = request.GET.get("group_type_id", None)
        if group_type_id:
            try:
                group_type = GroupType.objects.get(id=group_type_id)
            except GroupType.DoesNotExist:
                logger.warning(f"GroupType with id {group_type_id} does not exist")
                return Response({"results": []})

            queryset = queryset.filter(group_type=group_type)
            occurrence_report_id = request.GET.get("occurrence_report_id", None)
            if occurrence_report_id:
                try:
                    occurrence_report = OccurrenceReport.objects.get(
                        id=occurrence_report_id
                    )
                except OccurrenceReport.DoesNotExist:
                    logger.warning(
                        "OccurrenceReport with id {} does not exist".format(
                            occurrence_report_id
                        )
                    )
                    return Response({"results": []})

                if group_type.name in [
                    GroupType.GROUP_TYPE_FLORA,
                    GroupType.GROUP_TYPE_FAUNA,
                ]:
                    queryset = queryset.filter(species=occurrence_report.species)
                elif group_type.name == GroupType.GROUP_TYPE_COMMUNITY:
                    queryset = queryset.filter(community=occurrence_report.community)

        search_term = request.GET.get("term", None)
        if search_term:
            if occurrence_report_id:
                queryset = (
                    queryset.annotate(
                        display_name=Concat(
                            "occurrence_number",
                            Value(" - "),
                            "occurrence_name",
                            Value(" ("),
                            "group_type__name",
                            Value(")"),
                            output_field=CharField(),
                        ),
                    )
                    .filter(display_name__icontains=search_term)
                    .distinct()
                    .values("id", "display_name")[:10]
                )
                queryset = [
                    {"id": occurrence["id"], "text": occurrence["display_name"]}
                    for occurrence in queryset
                ]
            else:
                queryset = (
                    queryset.filter(occurrence_name__icontains=search_term)
                    .distinct()
                    .values("id", "occurrence_name")[:10]
                )

                queryset = [
                    {"id": occurrence["id"], "text": occurrence["occurrence_name"]}
                    for occurrence in queryset
                ]
        return Response({"results": queryset})

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def combine_occurrence_name_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):  # TODO group auth
            main_occurrence_id = request.GET.get("occurrence_id", None)

            if main_occurrence_id:
                try:
                    main_occurrence = self.get_queryset().get(id=main_occurrence_id)
                    queryset = (
                        self.get_queryset()
                        .exclude(id=main_occurrence_id)
                        .exclude(
                            processing_status=Occurrence.PROCESSING_STATUS_HISTORICAL,
                        )
                        .exclude(
                            processing_status=Occurrence.PROCESSING_STATUS_DISCARDED,
                        )
                        .filter(group_type=main_occurrence.group_type)
                    )

                    if main_occurrence.group_type.name in [
                        GroupType.GROUP_TYPE_FLORA,
                        GroupType.GROUP_TYPE_FAUNA,
                    ]:
                        # get species and all parents/children of those species
                        species_ids = get_all_related_species(
                            main_occurrence.species.id
                        )
                        queryset = queryset.filter(species_id__in=species_ids)

                    search_term = request.GET.get("term", None)
                    if search_term and main_occurrence_id:
                        queryset = (
                            queryset.annotate(
                                display_name=Concat(
                                    "occurrence_number",
                                    Value(" - "),
                                    "occurrence_name",
                                    Value(" ("),
                                    "group_type__name",
                                    Value(")"),
                                    output_field=CharField(),
                                ),
                            )
                            .filter(display_name__icontains=search_term)
                            .distinct()
                            .values(
                                "id",
                                "display_name",
                                "occurrence_number",
                                "occurrence_name",
                                "occurrence_source",
                                "wild_status",
                                "review_due_date",
                                "comment",
                            )[:10]
                        )
                        queryset = [
                            {
                                "id": occurrence["id"],
                                "text": occurrence["display_name"],
                                "occurrence_number": occurrence["occurrence_number"],
                                "occurrence_name": occurrence["occurrence_name"],
                                "occurrence_source": occurrence["occurrence_source"],
                                "wild_status": occurrence["wild_status"],
                                "review_due_date": occurrence["review_due_date"],
                                "comment": occurrence["comment"],
                            }
                            for occurrence in queryset
                        ]
                except Exception as e:
                    print(e)
                    queryset = self.get_queryset().none()
            else:
                queryset = self.get_queryset().none()

            return Response({"results": queryset})
        return Response()

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def combine_key_contacts_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):
            occ_ids = json.loads(request.POST.get("occurrence_ids"))
            contacts = OCCContactDetail.objects.filter(
                occurrence__id__in=occ_ids
            ).filter(visible=True)

            values_list = list(
                contacts.values(
                    "occurrence__occurrence_number",
                    "occurrence__id",
                    "id",
                    "contact_name",
                    "role",
                    "contact",
                    "organisation",
                    "notes",
                )
            )
            id_list = list(contacts.values_list("id", flat=True))

            return Response({"values_list": values_list, "id_list": id_list})
        return Response()

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def combine_documents_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):
            occ_ids = json.loads(request.POST.get("occurrence_ids"))
            documents = OccurrenceDocument.objects.filter(
                occurrence__id__in=occ_ids
            ).filter(visible=True)

            values_list = list(
                documents.values(
                    "occurrence__occurrence_number",
                    "occurrence__id",
                    "id",
                    "document_number",
                    "document_category__document_category_name",
                    "document_sub_category__document_sub_category_name",
                    "name",
                    "_file",
                    "description",
                    "uploaded_date",
                )
            )
            id_list = list(documents.values_list("id", flat=True))

            return Response({"values_list": values_list, "id_list": id_list})
        return Response()

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def combine_threats_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):
            occ_ids = json.loads(request.POST.get("occurrence_ids"))
            threats = OCCConservationThreat.objects.filter(
                occurrence__id__in=occ_ids
            ).filter(visible=True)

            values_list = list(
                threats.values(
                    "occurrence__occurrence_number",
                    "occurrence__id",
                    "id",
                    "threat_number",
                    "occurrence_report_threat__occurrence_report__occurrence_report_number",
                    "occurrence_report_threat__threat_number",
                    "threat_category__name",
                    "date_observed",
                    "threat_agent__name",
                    "current_impact__name",
                    "potential_impact__name",
                    "comment",
                )
            )
            id_list = list(threats.values_list("id", flat=True))

            return Response({"values_list": values_list, "id_list": id_list})
        return Response()
    
    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def combine_sites_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):
            occ_ids = json.loads(request.POST.get("occurrence_ids"))
            sites = OccurrenceSite.objects.filter(
                occurrence__id__in=occ_ids
            ).filter(visible=True)

            values_list = OccurrenceSiteSerializer(sites,context={"request": request}, many=True)
            id_list = list(sites.values_list("id", flat=True))

            return Response({"values_list": values_list.data, "id_list": id_list})
        return Response()

    @list_route(
        methods=[
            "POST",
        ],
        detail=False,
    )
    def combine_tenures_lookup(self, request, *args, **kwargs):
        if is_internal(self.request):
            occ_ids = json.loads(request.POST.get("occurrence_ids"))
            tenures = OccurrenceTenure.objects.filter(
                Q(occurrence_geometry__occurrence__id__in=occ_ids) | Q(historical_occurrence__in=occ_ids)
            )

            values_list = ListOccurrenceTenureSerializer(tenures,context={"request": request}, many=True)
            id_list = list(tenures.values_list("id", flat=True))

            return Response({"values_list": values_list.data, "id_list": id_list})
        return Response()

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.documents.all()
        if is_internal(self.request):
            qs = instance.documents.all()
        else:
            qs = instance.documents.none()
        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_internal(self.request):
            qs = instance.occ_threats.all()
        else:
            qs = instance.occ_threats.none()
        qs = qs.order_by("-date_observed")
        serializer = OCCConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_occurrence_reports(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports()
        if is_internal(self.request):
            related_reports = related_reports.all()
        else:
            related_reports = related_reports.none()

        serializer = ListInternalOccurrenceReportSerializer(
            related_reports, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_items(self, request, *args, **kwargs):
        instance = self.get_object()
        related_filter_type = request.GET.get("related_filter_type")
        related_items = instance.get_related_items(related_filter_type)
        serializer = RelatedItemsSerializer(related_items, many=True)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the Related Items dashboard filters"""
        related_type = Occurrence.RELATED_ITEM_CHOICES
        res_json = json.dumps(related_type)
        return HttpResponse(res_json, content_type="application/json")


class OccurrenceDocumentViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OccurrenceDocument.objects.none()
    serializer_class = OccurrenceDocumentSerializer

    def get_queryset(self):
        qs = OccurrenceDocument.objects.none()

        if is_internal(self.request):
            qs = OccurrenceDocument.objects.all().order_by("id")

        return qs

    def is_authorised_to_update(self, occurrence):
        if not is_occurrence_approver(self.request):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_DISCARD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_REINSTATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        serializer = SaveOccurrenceDocumentSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(no_revision=True)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_UPDATE_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOccurrenceDocumentSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        self.is_authorised_to_update(serializer.validated_data["occurrence"])
        instance = serializer.save(no_revision=True)
        instance.add_documents(request, no_revision=True)
        instance.uploaded_by = request.user.id
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_ADD_DOCUMENT.format(
                    instance.document_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        return Response(serializer.data)


class OCCConservationThreatFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):

        total_count = queryset.count()

        filter_threat_source = request.GET.get("filter_threat_source")
        if filter_threat_source and not filter_threat_source.lower() == "all":
            queryset = queryset.filter(
                (
                    Q(occurrence__occurrence_number=filter_threat_source)
                    & Q(occurrence_report_threat__occurrence_report=None)
                )
                | Q(
                    occurrence_report_threat__occurrence_report__occurrence_report_number=filter_threat_source
                )
            )

        filter_threat_category = request.GET.get("filter_threat_category")
        if filter_threat_category and not filter_threat_category.lower() == "all":
            queryset = queryset.filter(threat_category_id=filter_threat_category)

        filter_threat_current_impact = request.GET.get("filter_threat_current_impact")
        if (
            filter_threat_current_impact
            and not filter_threat_current_impact.lower() == "all"
        ):
            queryset = queryset.filter(current_impact=filter_threat_current_impact)

        filter_threat_potential_impact = request.GET.get(
            "filter_threat_potential_impact"
        )
        if (
            filter_threat_potential_impact
            and not filter_threat_potential_impact.lower() == "all"
        ):
            queryset = queryset.filter(potential_impact=filter_threat_potential_impact)

        filter_threat_status = request.GET.get("filter_threat_status")
        if filter_threat_status and not filter_threat_status.lower() == "all":
            if filter_threat_status == "active":
                queryset = queryset.filter(visible=True)
            elif filter_threat_status == "removed":
                queryset = queryset.filter(visible=False)

        def get_date(filter_date):
            date = request.GET.get(filter_date)
            if date:
                date = datetime.strptime(date, "%Y-%m-%d")
            return date

        filter_observed_from_date = get_date("filter_observed_from_date")
        if filter_observed_from_date:
            queryset = queryset.filter(date_observed__gte=filter_observed_from_date)

        filter_observed_to_date = get_date("filter_observed_to_date")
        if filter_observed_to_date:
            queryset = queryset.filter(date_observed__lte=filter_observed_to_date)

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)

        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OCCConservationThreatViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OCCConservationThreat.objects.none()
    serializer_class = OCCConservationThreatSerializer
    filter_backends = (OCCConservationThreatFilterBackend,)

    def get_queryset(self):
        qs = OCCConservationThreat.objects.none()

        if is_internal(self.request):
            qs = OCCConservationThreat.objects.all().order_by("id")

        return qs

    def is_authorised_to_update(self, occurrence):
        if not is_occurrence_approver(self.request):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_DISCARD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = True
        instance.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_REINSTATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        serializer = SaveOCCConservationThreatSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        serializer.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_UPDATE_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = SaveOCCConservationThreatSerializer(
            data=json.loads(request.data.get("data"))
        )
        validate_threat_request(request)
        serializer.is_valid(raise_exception=True)
        self.is_authorised_to_update(serializer.validated_data["occurrence"])
        instance = serializer.save(version_user=request.user)
        if instance.occurrence:
            instance.occurrence.log_user_action(
                OccurrenceUserAction.ACTION_ADD_THREAT.format(
                    instance.threat_number,
                    instance.occurrence.occurrence_number,
                ),
                request,
            )
        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class GetWildStatus(views.APIView):
    def get(self, request, format=None):
        search_term = request.GET.get("term", "")
        if search_term:
            data = WildStatus.objects.filter(name__icontains=search_term).values(
                "id", "name"
            )[:10]
            data_transform = [
                {"id": wild_status["id"], "text": wild_status["name"]}
                for wild_status in data
            ]
            return Response({"results": data_transform})
        return Response()


class OccurrenceViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin, DatumSearchMixin
):
    queryset = Occurrence.objects.none()
    serializer_class = OccurrenceSerializer
    lookup_field = "id"

    def get_queryset(self):
        qs = Occurrence.objects.all()
        if is_customer(self.request):
            qs = qs.filter(submitter=self.request.user.id)
        return qs

    def is_authorised_to_update(self):
        instance = self.get_object()
        if not (
            is_occurrence_approver(self.request)
            and (
                instance.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE
                or instance.processing_status == Occurrence.PROCESSING_STATUS_DRAFT
            )
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        group_type_id = GroupType.objects.get(id=request.data.get("group_type_id"))

        new_instance = Occurrence(
            submitter=request.user.id,
            group_type=group_type_id,
        )

        if not is_occurrence_approver(self.request):
            raise serializers.ValidationError(
                "User not authorised to create Occurrence"
            )

        new_instance.save(version_user=request.user)
        data = {"occurrence_id": new_instance.id}

        # create Location for new instance
        serializer = SaveOCCLocationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatComposition for new instance
        serializer = SaveOCCHabitatCompositionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create HabitatCondition for new instance
        serializer = SaveOCCHabitatConditionSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create FireHistory for new instance
        serializer = SaveOCCFireHistorySerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AssociatedSpecies for new instance
        serializer = SaveOCCAssociatedSpeciesSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create ObservationDetail for new instance
        serializer = SaveOCCObservationDetailSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create PlantCount for new instance
        serializer = SaveOCCPlantCountSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create AnimalObservation for new instance
        serializer = SaveOCCAnimalObservationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # create Identification for new instance
        serializer = SaveOCCIdentificationSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        serialized_obj = CreateOccurrenceSerializer(new_instance)
        return Response(serialized_obj.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def section_values(self, request, *args, **kwargs):

        section = request.GET.get("section")
        occ = self.get_object()
        res_json = {}

        if hasattr(occ, section):
            section_value = getattr(occ, section)
            section_fields = section_value._meta.get_fields()

            for i in section_fields:
                if (
                    i.name == "id"
                    or i.name == "occurrence"
                    or isinstance(i, models.ManyToOneRel)
                ):
                    continue

                # ensure many to many fields are assigned an appropriate __str__
                if isinstance(i, models.ManyToManyField):
                    sub_section_values = getattr(section_value, i.name)
                    res_json[i.name] = []
                    for j in sub_section_values.all():
                        if j.__str__():
                            res_json[i.name].append(j.__str__())
                        else:
                            res_json[i.name].append(j.id)

                elif isinstance(i, models.ForeignKey):
                    sub_section_value = getattr(section_value, i.name)
                    if sub_section_value is not None:
                        res_json[i.name] = {}
                        sub_section_fields = sub_section_value._meta.get_fields()
                        for j in sub_section_fields:
                            if (
                                j.name != "id"
                                and not isinstance(j, models.ForeignKey)
                                and not isinstance(j, models.ManyToOneRel)
                                and not isinstance(j, models.ManyToManyRel)
                                and getattr(sub_section_value, j.name) is not None
                            ):
                                res_json[i.name][j.name] = str(
                                    getattr(sub_section_value, j.name)
                                )
                        # if the num sub section has only one value, assign as section
                        if len(res_json[i.name]) == 1:
                            res_json[i.name] = list(res_json[i.name].values())[0]
                elif isinstance(i, MultiSelectField):
                    if i.choices:
                        choice_dict = dict(i.choices)
                        id_list = getattr(section_value, i.name)
                        values_list = []
                        for id in id_list:
                            if id.isdigit() and int(id) in choice_dict:
                                values_list.append(choice_dict[int(id)])
                        res_json[i.name] = values_list
                    else:
                        res_json[i.name] = getattr(section_value, i.name)

                elif getattr(section_value, i.name) is not None:
                    res_json[i.name] = str(getattr(section_value, i.name))

        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def combine(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        # print(json.loads(request.POST.get("data")))
        instance = self.get_object()
        occ_combine_data = json.loads(request.POST.get("data"))
        combine_occurrences = Occurrence.objects.exclude(id=instance.id).filter(
            id__in=occ_combine_data["combine_ids"]
        )
        # validate species
        if instance.group_type.name in [
            GroupType.GROUP_TYPE_FLORA,
            GroupType.GROUP_TYPE_FAUNA,
        ]:
            # get species and all parents/children of those species
            species_ids = get_all_related_species(instance.species.id)
            if (
                combine_occurrences.filter(species_id__in=species_ids).count()
                != combine_occurrences.count()
            ):
                raise serializers.ValidationError(
                    "Selected Occurrence has invalid Species"
                )

        instance.combine(request)

        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "PATCH",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.discard(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "PATCH",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.reinstate(request)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def activate(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        instance.activate(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def lock_occurrence(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        instance.lock(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def unlock_occurrence(self, request, *args, **kwargs):
        instance = self.get_object()
        if (
            not is_occurrence_approver(self.request)
            and instance.processing_status == Occurrence.PROCESSING_STATUS_LOCKED
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )
        instance.unlock(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def close_occurrence(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        instance.close(request)
        return redirect(reverse("internal"))

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def action_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.action_logs.all()
        serializer = OccurrenceUserActionSerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.comms_logs.all()
        serializer = OccurrenceLogEntrySerializer(qs, many=True)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def add_comms_log(self, request, *args, **kwargs):
        instance = self.get_object()
        mutable = request.data._mutable
        request.data._mutable = True
        request.data["occurrence"] = f"{instance.id}"
        request.data["staff"] = f"{request.user.id}"
        request.data._mutable = mutable
        serializer = OccurrenceLogEntrySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comms = serializer.save()
        # Save the files
        for f in request.FILES:
            document = comms.documents.create()
            document.name = str(request.FILES[f])
            document._file = request.FILES[f]
            document.save()
        # End Save Documents

        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def documents(self, request, *args, **kwargs):
        instance = self.get_object()
        # qs = instance.documents.all()
        if is_internal(self.request):
            qs = instance.documents.all()
        else:
            qs = instance.documents.none()
        qs = qs.order_by("-uploaded_date")
        serializer = OccurrenceDocumentSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def threats(self, request, *args, **kwargs):
        instance = self.get_object()
        if is_internal(self.request):
            qs = instance.occ_threats.all()
        else:
            qs = instance.occ_threats.none()
        filter_backend = OCCConservationThreatFilterBackend()
        qs = filter_backend.filter_queryset(self.request, qs, self)
        serializer = OCCConservationThreatSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    # gets all distinct threat sources for threats pertaining to a specific OCC
    def threat_source_list(self, request, *args, **kwargs):
        instance = self.get_object()
        data = []
        if is_internal(self.request):
            # distinct on OCR
            qs = instance.occ_threats.distinct(
                "occurrence_report_threat__occurrence_report"
            ).exclude(occurrence_report_threat=None)
            # format
            data = [
                threat.occurrence_report_threat.occurrence_report.occurrence_report_number
                for threat in qs
            ]

        # if any occ threats exist with an ocr threat, then the source must be the occ
        if instance.occ_threats.filter(occurrence_report_threat=None).exists():
            data.append(instance.occurrence_number)

        return Response(data)

    @detail_route(methods=["get"], detail=True)
    def add_related_species(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        if hasattr(instance, "associated_species"):
            related_species = instance.associated_species.related_species
        else:
            raise serializers.ValidationError("Associated Species does not exist")

        taxon_id = request.GET.get("species")

        try:
            taxon = Taxonomy.objects.get(id=taxon_id)
        except Taxonomy.DoesNotExist:
            raise serializers.ValidationError("Species does not exist")

        if taxon not in related_species.all():
            related_species.add(taxon)

        instance.save(version_user=request.user)

        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def remove_related_species(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        if hasattr(instance, "associated_species"):
            related_species = instance.associated_species.related_species
        else:
            raise serializers.ValidationError("Associated Species does not exist")

        taxon_id = request.GET.get("species")

        try:
            taxon = Taxonomy.objects.get(id=taxon_id)
        except Taxonomy.DoesNotExist:
            raise serializers.ValidationError("Species does not exist")

        if taxon in related_species.all():
            related_species.remove(taxon)

        instance.save(version_user=request.user)

        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_species(self, request, *args, **kwargs):
        instance = self.get_object()
        if hasattr(instance, "associated_species"):
            related_species = instance.associated_species.related_species
        else:
            related_species = Taxonomy.objects.none()
        serializer = TaxonomySerializer(
            related_species, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_occurrence_reports(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports()
        if is_internal(self.request):
            related_reports = related_reports.all()
        else:
            related_reports = related_reports.none()
        serializer = ListInternalOccurrenceReportSerializer(
            related_reports, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_existing_ocr_threats(self, request, *args, **kwargs):
        instance = self.get_object()
        related_reports = instance.get_related_occurrence_reports().values_list(
            "id", flat=True
        )
        addedThreats = (
            OCCConservationThreat.objects.filter(occurrence=instance)
            .exclude(occurrence_report_threat=None)
            .values_list("occurrence_report_threat_id", flat=True)
        )
        threats = OCRConservationThreat.objects.filter(
            occurrence_report_id__in=related_reports
        ).exclude(id__in=addedThreats)
        if is_internal(self.request):
            threats = threats.all()
        else:
            threats = threats.none()
        serializer = OCRConservationThreatSerializer(
            threats, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["get"], detail=True)
    def get_related_items(self, request, *args, **kwargs):
        instance = self.get_object()
        related_filter_type = request.GET.get("related_filter_type")
        related_items = instance.get_related_items(related_filter_type)
        serializer = RelatedItemsSerializer(related_items, many=True)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def filter_list(self, request, *args, **kwargs):
        """Used by the Related Items dashboard filters"""
        related_type = Occurrence.RELATED_ITEM_CHOICES
        res_json = json.dumps(related_type)
        return HttpResponse(res_json, content_type="application/json")

    # used for Location Tab of Occurrence external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="location-list-of-values",
    )
    def location_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence external form"""
        qs = self.get_queryset()
        datum_list = []

        id = request.GET.get("id", None)
        try:
            qs = qs.get(id=id)
        except Occurrence.DoesNotExist:
            logger.error(f"Occurrence with id {id} not found")
        else:
            pass
            occ_geometries = qs.occ_geometry.all().exclude(**{"geometry": None})
            epsg_codes = [
                str(g.srid)
                for g in occ_geometries.values_list("geometry", flat=True).distinct()
            ]
            # Add the srids of the original geometries to epsg_codes
            original_geometry_srids = [
                str(g.original_geometry_srid) for g in occ_geometries
            ]
            epsg_codes += [g for g in original_geometry_srids if g.isnumeric()]
            epsg_codes = list(set(epsg_codes))
            datum_list = search_datums("", codes=epsg_codes)

        coordinate_source_list = []
        values = CoordinateSource.objects.all()
        if values:
            for val in values:
                coordinate_source_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        location_accuracy_list = []
        values = LocationAccuracy.objects.all()
        if values:
            for val in values:
                location_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "datum_list": datum_list,
            "coordinate_source_list": coordinate_source_list,
            "location_accuracy_list": location_accuracy_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def occurrence_save(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        instance = self.get_object()
        request_data = request.data

        if request_data.get("habitat_composition"):
            habitat_instance, created = OCCHabitatComposition.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCHabitatCompositionSerializer(
                habitat_instance, data=request_data.get("habitat_composition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("habitat_condition"):
            hab_cond_instance, created = OCCHabitatCondition.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCHabitatConditionSerializer(
                hab_cond_instance, data=request_data.get("habitat_condition")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("vegetation_structure"):
            veg_struct_instance, created = OCCVegetationStructure.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCVegetationStructureSerializer(
                veg_struct_instance, data=request_data.get("vegetation_structure")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("fire_history"):
            fire_instance, created = OCCFireHistory.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCFireHistorySerializer(
                fire_instance, data=request_data.get("fire_history")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("associated_species"):
            assoc_species_instance, created = (
                OCCAssociatedSpecies.objects.get_or_create(occurrence=instance)
            )
            serializer = SaveOCCAssociatedSpeciesSerializer(
                assoc_species_instance,
                data=request_data.get("associated_species"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("observation_detail"):
            obs_det_instance, created = OCCObservationDetail.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCObservationDetailSerializer(
                obs_det_instance, data=request_data.get("observation_detail")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("plant_count"):
            plant_count_instance, created = OCCPlantCount.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCPlantCountSerializer(
                plant_count_instance, data=request_data.get("plant_count")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("animal_observation"):
            animal_obs_instance, created = OCCAnimalObservation.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCAnimalObservationSerializer(
                animal_obs_instance,
                data=request_data.get("animal_observation"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("identification"):
            identification_instance, created = OCCIdentification.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCIdentificationSerializer(
                identification_instance,
                data=request_data.get("identification"),
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        if request_data.get("location"):
            location_instance, created = OCCLocation.objects.get_or_create(
                occurrence=instance
            )
            serializer = SaveOCCLocationSerializer(
                location_instance, data=request_data.get("location")
            )
            serializer.is_valid(raise_exception=True)
            if serializer.is_valid():
                serializer.save()

        # occ geometry data to save seperately
        geometry_data = request_data.get("occ_geometry", None)
        if geometry_data:
            intersect_data = save_geometry(
                request, instance, geometry_data, "occurrence"
            )
            instance.occ_geometry.all()

            if intersect_data:
                for key, value in intersect_data.items():
                    occurrence_geometry = OccurrenceGeometry.objects.get(id=key)
                    populate_occurrence_tenure_data(
                        occurrence_geometry, value.get("features", [])
                    )

        occ_sites = OccurrenceSite.objects
        site_geometry_data = json.loads(request.data.get("site_geometry", None))
        if site_geometry_data and "features" in site_geometry_data:
            for i in site_geometry_data["features"]:
                try:
                    update_site = occ_sites.get(site_number=i["properties"]["site_number"])
                    point_data = 'POINT({0} {1})'.format(i["geometry"]["coordinates"][0],i["geometry"]["coordinates"][1])
                    new_geometry = GEOSGeometry(point_data, srid=i["properties"]["srid"])
                    update_site.geometry = new_geometry
                    update_site.save() #TODO add version_user when history implemented
                except Exception as e:
                    print(e)

        serializer = SaveOccurrenceSerializer(instance, data=request_data, partial=True)
        serializer.is_valid(raise_exception=True)

        if serializer.is_valid():
            serializer.save(version_user=request.user)

            instance.log_user_action(
                OccurrenceUserAction.ACTION_SAVE_OCCURRENCE.format(
                    instance.occurrence_number
                ),
                request,
            )

        final_instance = self.get_object()
        serializer = self.get_serializer(final_instance)
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @transaction.atomic
    def copy_ocr_section(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        instance = self.get_object()
        data = json.loads(request.data["data"])

        ocrId = data["occurrence_report_id"]
        section = data["section"]

        ocr = OccurrenceReport.objects.get(id=ocrId)
        ocrSection = getattr(ocr, section)
        occSection = getattr(instance, section)

        section_fields = type(ocrSection)._meta.get_fields()
        for i in section_fields:
            if (
                i.name != "id"
                and i.name != "occurrence_report"
                and hasattr(occSection, i.name)
            ):
                if isinstance(i, models.ManyToManyField):
                    ocrValue = getattr(ocrSection, i.name)
                    occValue = getattr(occSection, i.name)
                    occValue.clear()
                    for i in ocrValue.all():
                        occValue.add(i)
                else:
                    ocrValue = getattr(ocrSection, i.name)
                    setattr(occSection, i.name, ocrValue)

        occ_section_fields = type(occSection)._meta.get_fields()
        for i in occ_section_fields:
            if (
                isinstance(i, models.ForeignKey)
                and i.related_model.__name__ == ocrSection.__class__.__name__
            ):
                setattr(occSection, i.name, ocrSection)

        occSection.save()
        instance.save(version_user=request.user)

        serialized_obj = OccurrenceSerializer(instance, context={"request": request})
        return Response(serialized_obj.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_location_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()

        location_instance, created = OCCLocation.objects.get_or_create(
            occurrence=occ_instance
        )

        # occ geometry data to save seperately
        geometry_data = request.data.get("occ_geometry")
        if geometry_data:
            intersect_data = save_geometry(
                request, occ_instance, geometry_data, "occurrence"
            )
            if intersect_data:
                for key, value in intersect_data.items():
                    occurrence_geometry = OccurrenceGeometry.objects.get(id=key)
                    populate_occurrence_tenure_data(
                        occurrence_geometry, value.get("features", [])
                    )

        occ_sites = OccurrenceSite.objects
        site_geometry_data = json.loads(request.data.get("site_geometry", None))
        if site_geometry_data and "features" in site_geometry_data:
            for i in site_geometry_data["features"]:
                try:
                    update_site = occ_sites.get(site_number=i["properties"]["site_number"])
                    point_data = 'POINT({0} {1})'.format(i["geometry"]["coordinates"][0],i["geometry"]["coordinates"][1])
                    new_geometry = GEOSGeometry(point_data, srid=i["properties"]["srid"])
                    update_site.geometry = new_geometry
                    update_site.save() #TODO add version_user when history implemented
                except Exception as e:
                    print(e)

        # the request.data is only the habitat composition data thats been sent from front end
        location_data = request.data.get("location")
        serializer = SaveOCCLocationSerializer(
            location_instance, data=location_data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_composition_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        habitat_instance, created = OCCHabitatComposition.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCHabitatCompositionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_vegetation_structure(self, request, *args, **kwargs):

        self.is_authorised_to_update()
        occ_instance = self.get_object()
        vegetation_instance, created = OCCVegetationStructure.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCCVegetationStructureSerializer(
            vegetation_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_habitat_condition_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        habitat_instance, created = OCCHabitatCondition.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat condition data thats been sent from front end
        serializer = SaveOCCHabitatConditionSerializer(
            habitat_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_fire_history_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        fire_instance, created = OCCFireHistory.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCFireHistorySerializer(
            fire_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_associated_species_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        assoc_species_instance, created = OCCAssociatedSpecies.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the habitat composition data thats been sent from front end
        serializer = SaveOCCAssociatedSpeciesSerializer(
            assoc_species_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_observation_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        obs_det_instance, created = OCCObservationDetail.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the observation detail data thats been sent from front end
        serializer = SaveOCCObservationDetailSerializer(
            obs_det_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_plant_count_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        plant_count_instance, created = OCCPlantCount.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the plant count data thats been sent from front end
        serializer = SaveOCCPlantCountSerializer(
            plant_count_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_animal_observation_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        animal_obs_instance, created = OCCAnimalObservation.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the animal obs data thats been sent from front end
        serializer = SaveOCCAnimalObservationSerializer(
            animal_obs_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @list_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def update_identification_details(self, request, *args, **kwargs):
        self.is_authorised_to_update()
        occ_instance = self.get_object()
        identification_instance, created = OCCIdentification.objects.get_or_create(
            occurrence=occ_instance
        )
        # the request.data is only the identification data thats been sent from front end
        serializer = SaveOCCIdentificationSerializer(
            identification_instance, data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    # used for Occurrence external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def list_of_values(self, request, *args, **kwargs):
        """used for Occurrence external form"""
        land_form_list = []
        types = LandForm.objects.all()
        if types:
            for val in types:
                land_form_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        rock_type_list = []
        types = RockType.objects.all()
        if types:
            for val in types:
                rock_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_type_list = []
        types = SoilType.objects.all()
        if types:
            for val in types:
                soil_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_colour_list = []
        colours = SoilColour.objects.all()
        if colours:
            for val in colours:
                soil_colour_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        soil_condition_list = []
        conditions = SoilCondition.objects.all()
        if conditions:
            for val in conditions:
                soil_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        drainage_list = []
        drainages = Drainage.objects.all()
        if drainages:
            for val in drainages:
                drainage_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        intensity_list = []
        intensities = Intensity.objects.all()
        if intensities:
            for val in intensities:
                intensity_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "land_form_list": land_form_list,
            "rock_type_list": rock_type_list,
            "soil_type_list": soil_type_list,
            "soil_colour_list": soil_colour_list,
            "soil_condition_list": soil_condition_list,
            "drainage_list": drainage_list,
            "intensity_list": intensity_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    # used for Occurrence Observation external form
    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def observation_list_of_values(self, request, *args, **kwargs):
        """used for Occurrence external form"""
        observation_method_list = []
        values = ObservationMethod.objects.all()
        if values:
            for val in values:
                observation_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_method_list = []
        values = PlantCountMethod.objects.all()
        if values:
            for val in values:
                plant_count_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_count_accuracy_list = []
        values = PlantCountAccuracy.objects.all()
        if values:
            for val in values:
                plant_count_accuracy_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        plant_condition_list = []
        values = PlantCondition.objects.all()
        if values:
            for val in values:
                plant_condition_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        counted_subject_list = []
        values = CountedSubject.objects.all()
        if values:
            for val in values:
                counted_subject_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        primary_detection_method_list = []
        values = PrimaryDetectionMethod.objects.all()
        if values:
            for val in values:
                primary_detection_method_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        secondary_sign_list = []
        values = SecondarySign.objects.all()
        if values:
            for val in values:
                secondary_sign_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        reprod_state_list = []
        values = ReproductiveState.objects.all()
        if values:
            for val in values:
                reprod_state_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        death_reason_list = []
        values = DeathReason.objects.all()
        if values:
            for val in values:
                death_reason_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        animal_health_list = []
        values = AnimalHealth.objects.all()
        if values:
            for val in values:
                animal_health_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        identification_certainty_list = []
        values = IdentificationCertainty.objects.all()
        if values:
            for val in values:
                identification_certainty_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_type_list = []
        values = SampleType.objects.all()
        if values:
            for val in values:
                sample_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        sample_dest_list = []
        values = SampleDestination.objects.all()
        if values:
            for val in values:
                sample_dest_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        permit_type_list = []
        values = PermitType.objects.all()
        if values:
            for val in values:
                permit_type_list.append(
                    {
                        "id": val.id,
                        "name": val.name,
                    }
                )
        res_json = {
            "observation_method_list": observation_method_list,
            "plant_count_method_list": plant_count_method_list,
            "plant_count_accuracy_list": plant_count_accuracy_list,
            "plant_condition_list": plant_condition_list,
            "counted_subject_list": counted_subject_list,
            "primary_detection_method_list": primary_detection_method_list,
            "secondary_sign_list": secondary_sign_list,
            "reprod_state_list": reprod_state_list,
            "death_reason_list": death_reason_list,
            "animal_health_list": animal_health_list,
            "identification_certainty_list": identification_certainty_list,
            "sample_type_list": sample_type_list,
            "sample_dest_list": sample_dest_list,
            "permit_type_list": permit_type_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
        url_path="available-occurrence-reports-crs",
    )
    def available_occurrence_reports_crs(self, request, *args, **kwargs):
        """used for Occurrence Report external form"""
        qs = self.get_queryset()
        crs = []

        id = request.GET.get("id", None)
        try:
            qs = qs.get(id=id)
        except Occurrence.DoesNotExist:
            logger.error(f"Occurrence with id {id} not found")
        else:
            ocr_geometries_ids = (
                qs.occurrence_reports.all()
                .values_list("ocr_geometry", flat=True)
                .distinct()
            )
            ocr_geometries = OccurrenceReportGeometry.objects.filter(
                id__in=ocr_geometries_ids
            ).exclude(**{"geometry": None})

            epsg_codes = [
                str(g.srid)
                for g in ocr_geometries.values_list("geometry", flat=True).distinct()
            ]
            # Add the srids of the original geometries to epsg_codes
            original_geometry_srids = [
                str(g.original_geometry_srid) for g in ocr_geometries
            ]
            epsg_codes += [g for g in original_geometry_srids if g.isnumeric()]
            epsg_codes = list(set(epsg_codes))
            crs = search_datums("", codes=epsg_codes)

        res_json = {
            "crs": crs,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")

    @list_route(methods=["GET"], detail=False)
    def list_for_map(self, request, *args, **kwargs):
        occurrence_ids = [
            int(id)
            for id in request.query_params.get("proposal_ids", "").split(",")
            if id.lstrip("-").isnumeric()
        ]

        cache_key = settings.CACHE_KEY_MAP_OCCURRENCES
        qs = cache.get(cache_key)
        if qs is None:
            qs = (
                self.get_queryset()
                .exclude(occ_geometry__isnull=True)
                .prefetch_related("occ_geometry")
            )
            cache.set(cache_key, qs, settings.CACHE_TIMEOUT_2_HOURS)

        if len(occurrence_ids) > 0:
            qs = qs.filter(id__in=occurrence_ids)

        serializer = ListOCCMinimalSerializer(
            qs, context={"request": request}, many=True
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def contact_details(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.contact_detail.all()
        serializer = OCCContactDetailSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def sites(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = instance.sites.all()
        serializer = OccurrenceSiteSerializer(
            qs, many=True, context={"request": request}
        )
        return Response(serializer.data)


class OccurrenceReportReferralViewSet(
    viewsets.GenericViewSet, mixins.RetrieveModelMixin
):
    queryset = OccurrenceReportReferral.objects.all()
    serializer_class = OccurrenceReportReferralSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if not is_internal(self.request):
            if is_contributor(self.request) and is_occurrence_report_referee(
                self.request
            ):
                qs = qs.filter(
                    Q(occurrence_report__submitter=self.request.user.id)
                    | Q(referral=self.request.user.id)
                )
            elif is_contributor(self.request):
                qs = qs.filter(occurrence_report__submitter=self.request.user.id)
            elif is_occurrence_report_referee(self.request):
                qs = qs.filter(referral=self.request.user.id)
        return qs

    def is_authorised_to_refer(self):
        instance = self.get_object()
        if not instance.occurrence_report.has_assessor_mode(self.request):
            raise serializers.ValidationError(
                "User not authorised to manage Referrals for Occurrence Report"
            )

    def is_authorised_to_referee(self):
        instance = self.get_object()
        user = self.request.user
        if not instance.referral == user.id:
            raise serializers.ValidationError(
                "User is not the Referee for Occurrence Report Referral"
            )

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def referral_list(self, request, *args, **kwargs):
        instance = self.get_object()
        qs = self.get_queryset().filter(
            sent_by=instance.referral, occurrence_report=instance.occurrence_report
        )
        serializer = self.get_serializer(qs, many=True, context={"request": request})
        return Response(serializer.data)

    @detail_route(methods=["GET", "POST"], detail=True)
    def complete(self, request, *args, **kwargs):

        self.is_authorised_to_referee()

        instance = self.get_object()
        instance.complete(request)
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def remind(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.remind(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def recall(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.recall(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(
        methods=[
            "GET",
        ],
        detail=True,
    )
    def resend(self, request, *args, **kwargs):

        self.is_authorised_to_refer()

        instance = self.get_object()
        instance.resend(request)
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data)

    @detail_route(methods=["post"], detail=True)
    @renderer_classes((JSONRenderer,))
    @transaction.atomic
    def occurrence_report_referral_save(self, request, *args, **kwargs):

        self.is_authorised_to_referee()

        instance = self.get_object()
        request_data = request.data
        instance.referral_comment = request_data.get("referral_comment")
        instance.save()

        # Create a log entry for the occurrence report
        instance.occurrence_report.log_user_action(
            OccurrenceReportUserAction.COMMENT_REFERRAL.format(
                instance.id,
                instance.occurrence_report.occurrence_report_number,
                f"{instance.referral_as_email_user.get_full_name()}({instance.referral_as_email_user.email})",
            ),
            request,
        )
        serializer = self.get_serializer(instance, context={"request": request})
        return Response(serializer.data)


class OccurrenceTenureFilterBackend(DatatablesFilterBackend):
    def filter_queryset(self, request, queryset, view):
        total_count = queryset.count()

        query_params = {
            p: request.query_params[p]
            for p in request.query_params
            if request.query_params[p] not in ["all"]
        }

        filter_status = query_params.get("filter_status", None)
        tenure_area_id = query_params.get("tenure_area_id", None)
        vesting = query_params.get("vesting", None)
        purpose = query_params.get("purpose", None)

        queryset = queryset.filter(status=filter_status) if filter_status else queryset
        queryset = (
            queryset.filter(tenure_area_id=tenure_area_id)
            if tenure_area_id
            else queryset
        )
        # TODO: Implement vesting filtering after implementing the vesting field
        logger.debug(f"vesting: {vesting}")  # use variable or remove
        # queryset = queryset.filter(vesting=vesting) if vesting else queryset
        queryset = queryset.filter(purpose=purpose) if purpose else queryset

        fields = self.get_fields(request)
        ordering = self.get_ordering(request, view, fields)
        queryset = queryset.order_by(*ordering)
        if len(ordering):
            queryset = queryset.order_by(*ordering)

        queryset = super().filter_queryset(request, queryset, view)

        setattr(view, "_datatables_total_count", total_count)
        return queryset


class OccurrenceTenurePaginatedViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = OccurrenceTenure.objects.none()
    serializer_class = OccurrenceTenureSerializer
    pagination_class = DatatablesPageNumberPagination
    filter_backends = [
        OccurrenceTenureFilterBackend,
    ]
    page_size = 10

    def get_serializer_class(self):
        if self.action in ["list", "occurrence_tenure_internal"]:
            return ListOccurrenceTenureSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = super().get_queryset()
        if not is_internal(self.request):
            return qs.none()
        occurrence_id = self.request.query_params.get("occurrence_id", None)
        if occurrence_id and occurrence_id.isnumeric():
            return OccurrenceTenure.objects.filter(
                Q(occurrence_geometry__occurrence_id=occurrence_id)
                | Q(historical_occurrence=occurrence_id)
            )
        return OccurrenceTenure.objects.all()

    def current_and_historical_tenures(self, queryset, occurrence_id):
        return queryset.filter(
            models.Q(occurrence_geometry__occurrence_id=occurrence_id)
            | models.Q(
                ("historical_occurrence", occurrence_id),
                ("status", OccurrenceTenure.STATUS_HISTORICAL),
                _connector=models.Q.AND,
            )
        )

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_tenure_internal(self, request, *args, **kwargs):
        qs = self.get_queryset()
        qs = self.filter_queryset(qs)

        self.paginator.page_size = qs.count()
        result_page = self.paginator.paginate_queryset(qs, request)
        Serializer = self.get_serializer_class()
        serializer = Serializer(result_page, context={"request": request}, many=True)

        return self.paginator.get_paginated_response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_tenure_feature_id_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        search_term = request.GET.get("term", "")
        occurrence_id = request.GET.get("occurrence_id", None)

        if occurrence_id:
            queryset = self.current_and_historical_tenures(queryset, occurrence_id)

        if search_term:
            feature_id = models.Func(
                models.F("tenure_area_id"),
                Value("([0-9]+$)"),
                function="substring",
                output=models.TextField(),
            )
            queryset = queryset.annotate(feature_id=feature_id)

            queryset = (
                queryset.filter(feature_id__icontains=search_term)
                .distinct()
                .values("tenure_area_id", "feature_id")[:10]
            )
            results = [
                {"id": row["tenure_area_id"], "text": row["feature_id"]}
                for row in queryset
            ]

        return Response({"results": results})

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_tenure_vesting_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        search_term = request.GET.get("term", "")
        occurrence_id = request.GET.get("occurrence_id", None)

        if occurrence_id:
            queryset = self.current_and_historical_tenures(queryset, occurrence_id)

        results = []
        if search_term:
            # TODO: Implement vesting filtering after implementing the vesting field
            # queryset = queryset.filter(vesting__icontains=search_term).distinct()[:10]
            # results = [
            #     {"id": row.vesting, "text": row.vesting} for row in queryset
            # ]
            results = [{"id": 1, "text": queryset[0].vesting}]

        return Response({"results": results})

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def occurrence_tenure_purpose_lookup(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        search_term = request.GET.get("term", "")
        occurrence_id = request.GET.get("occurrence_id", None)

        if occurrence_id:
            queryset = self.current_and_historical_tenures(queryset, occurrence_id)

        if search_term:
            queryset = queryset.filter(
                purpose__purpose__icontains=search_term
            ).distinct()[:10]
            results = [
                {"id": row.purpose.id, "text": row.purpose.purpose} for row in queryset
            ]

        return Response({"results": results})


class ContactDetailViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OCCContactDetail.objects.none()
    serializer_class = OCCContactDetailSerializer

    def is_authorised_to_update(self, occurrence):
        if not (
            is_occurrence_approver(self.request)
            and (
                occurrence.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE
                or occurrence.processing_status == Occurrence.PROCESSING_STATUS_DRAFT
            )
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )

    def get_queryset(self):
        qs = OCCContactDetail.objects.none()

        if is_internal(self.request):
            qs = OCCContactDetail.objects.all().order_by("id")
        return qs

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        if not instance.visible:
            raise serializers.ValidationError("Discarded contact cannot be updated.")

        self.is_authorised_to_update(instance.occurrence)
        serializer = OCCContactDetailSerializer(
            instance, data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)

        occurrence = serializer.validated_data["occurrence"]
        contact_name = serializer.validated_data["contact_name"]

        if (
            OCCContactDetail.objects.exclude(id=instance.id)
            .filter(
                Q(contact_name=contact_name)
                & Q(occurrence=occurrence)
                & Q(visible=True)
            )
            .exists()
        ):
            raise serializers.ValidationError(
                "Contact with this name already exists for this occurrence"
            )

        serializer.save()

        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = OCCContactDetailSerializer(
            data=json.loads(request.data.get("data"))
        )
        serializer.is_valid(raise_exception=True)
        occurrence = serializer.validated_data["occurrence"]
        contact_name = serializer.validated_data["contact_name"]

        if OCCContactDetail.objects.filter(
            Q(contact_name=contact_name) & Q(occurrence=occurrence) & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Contact with this name already exists for this occurrence"
            )

        self.is_authorised_to_update(occurrence)
        serializer.save()

        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)

        if OCCContactDetail.objects.filter(
            Q(contact_name=instance.contact_name)
            & Q(occurrence=instance.occurrence)
            & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Active contact with this name already exists for this occurrence"
            )

        instance.visible = True
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)


class OccurrenceSiteViewSet(viewsets.GenericViewSet, mixins.RetrieveModelMixin):
    queryset = OccurrenceSite.objects.none()
    serializer_class = OccurrenceSiteSerializer

    def is_authorised_to_update(self, occurrence):
        if not (
            is_occurrence_approver(self.request)
            and (
                occurrence.processing_status == Occurrence.PROCESSING_STATUS_ACTIVE
                or occurrence.processing_status == Occurrence.PROCESSING_STATUS_DRAFT
            )
        ):
            raise serializers.ValidationError(
                "User not authorised to update Occurrence"
            )

    def get_queryset(self):
        qs = OccurrenceSite.objects.none()

        if is_internal(self.request):
            qs = OccurrenceSite.objects.all().order_by("id")
        return qs

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        if not instance.visible:
            raise serializers.ValidationError("Discarded site cannot be updated.")

        self.is_authorised_to_update(instance.occurrence)

        data = json.loads(request.data.get("data"))
        point_data = 'POINT({0} {1})'.format(data["point_coord1"],data["point_coord2"])
        data["geometry"] = GEOSGeometry(point_data, srid=data["datum"])

        serializer = SaveOccurrenceSiteSerializer(
            instance, data=data
        )
        serializer.is_valid(raise_exception=True)

        occurrence = serializer.validated_data["occurrence"]
        site_name = serializer.validated_data["site_name"]

        if (
            OccurrenceSite.objects.exclude(id=instance.id)
            .filter(Q(site_name=site_name) & Q(occurrence=occurrence) & Q(visible=True))
            .exists()
        ):
            raise serializers.ValidationError(
                "Site with this name already exists for this occurrence"
            )

        instance = serializer.save()

        return Response(serializer.data)

    def create(self, request, *args, **kwargs):

        data = json.loads(request.data.get("data"))
        point_data = 'POINT({0} {1})'.format(data["point_coord1"],data["point_coord2"])
        data["geometry"] = GEOSGeometry(point_data, srid=data["datum"])

        serializer = SaveOccurrenceSiteSerializer(
            data=data
        )
        serializer.is_valid(raise_exception=True)
        occurrence = serializer.validated_data["occurrence"]
        site_name = serializer.validated_data["site_name"]

        if OccurrenceSite.objects.filter(
            Q(site_name=site_name) & Q(occurrence=occurrence) & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Site with this name already exists for this occurrence"
            )

        self.is_authorised_to_update(occurrence)
        serializer.save()

        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def discard(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)
        instance.visible = False
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @detail_route(
        methods=[
            "POST",
        ],
        detail=True,
    )
    def reinstate(self, request, *args, **kwargs):
        instance = self.get_object()
        self.is_authorised_to_update(instance.occurrence)

        if OccurrenceSite.objects.filter(
            Q(site_name=instance.site_name)
            & Q(occurrence=instance.occurrence)
            & Q(visible=True)
        ).exists():
            raise serializers.ValidationError(
                "Active site with this name already exists for this occurrence"
            )

        instance.visible = True
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @list_route(
        methods=[
            "GET",
        ],
        detail=False,
    )
    def site_list_of_values(self, request, *args, **kwargs):

        site_type_list = list(SiteType.objects.values("id", "name"))
        datum_list = list({"srid":datum.srid,"name":datum.name} for datum in Datum.objects.all())

        res_json = {
            "site_type_list": site_type_list,
            "datum_list": datum_list,
        }
        res_json = json.dumps(res_json)
        return HttpResponse(res_json, content_type="application/json")
    
    @list_route(methods=["GET"], detail=False)
    def list_for_map(self, request, *args, **kwargs):
        occurrence_id = request.GET.get("occurrence_id")
        print(occurrence_id)
        qs = self.get_queryset().filter(occurrence_id=occurrence_id).exclude(geometry=None).exclude(visible=False)
        print(qs.count())
        serializer = SiteGeometrySerializer(
            qs, many=True
        )
        print(serializer.data)
        return Response(serializer.data)


class OCRExternalRefereeInviteViewSet(viewsets.ModelViewSet):
    queryset = OCRExternalRefereeInvite.objects.filter(archived=False)
    serializer_class = OCRExternalRefereeInviteSerializer

    def get_queryset(self):
        qs = self.queryset
        if not is_occurrence_assessor(self.request):
            qs = OCRExternalRefereeInvite.objects.none()
        return qs

    @detail_route(methods=["post"], detail=True)
    def remind(self, request, *args, **kwargs):
        instance = self.get_object()
        send_external_referee_invite_email(
            instance.occurrence_report, request, instance, reminder=True
        )
        return Response(
            status=status.HTTP_200_OK,
            data={"message": f"Reminder sent to {instance.email} successfully"},
        )

    @detail_route(methods=["patch"], detail=True)
    def retract(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.archived = True
        instance.save()
        serializer = InternalOccurrenceReportSerializer(
            instance.occurrence_report, context={"request": request}
        )
        return Response(serializer.data, status=status.HTTP_200_OK)
